from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from django.template.loader import render_to_string, get_template
from django.http import HttpResponse, FileResponse
from django.conf import settings
from django.utils.timezone import now, localdate
from django.db import transaction, connection, IntegrityError
from django.db.models import Max, Sum, Count, Min, Q, Prefetch, Exists, OuterRef, Subquery, DecimalField, IntegerField, Avg
from django.db.models.functions import Coalesce, Cast

from django_q.tasks import async_task
from django_tenants.utils import schema_context

from rest_framework.views import APIView
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from rest_framework.exceptions import ValidationError
from rest_framework.generics import ListAPIView

from datetime import datetime, date, timedelta
from weasyprint import HTML
from collections import defaultdict
from babel.dates import format_date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta

from apps.agua.core.permissions import GlobalPermissionMixin, TenantPaymentCreatePermission
from apps.tenant.utils.seed import generate_ticket
from apps.tenant.models import Pay, ReceiptBatch
from apps.user.models import User

from .models import Customer, ServiceRefinancingDetail, Manzana, CashMovement, ServiceCharge, Manzana, MeterAssignment, ServiceCut, Config, CutBatch, DailyCashReport, DebtRefinancing, DebtRefinancingDetail, RefinancingInstallment, WaterMeter, CashOutflow, CashBox, Reading, DebtDetail, CashConcept, Invoice, Category, Via, Calle, InvoiceDebt, InvoicePayment, Zona, Debt, ReadingGeneration, Company
from .serializers import (
    CustomerSerializer, ServiceCutSerializer, ServiceChargeSerializer, DebtRefinancingSerializer, ManzanaSerializer, MeterAssignmentSerializer, MorosidadSerializer, CutBatchSerializer, WaterMeterSerializer, RefinancingInstallmentSerializer, ViaSerializer, CompanySerializer, CashOutflowSerializer, CalleSerializer, DebtSerializer, CashBoxSerializer, CustomerWithDebtsSerializer,
    ReadingSerializer,  InvoiceSerializer, CategorySerializer, ZonaSerializer, ConfigSerializer, ReadingGenerationSerializer, CashConceptSerializer, DailyCashReportSerializer)

from .filters import ReadingFilter, DebtFilter
from .utils import get_catastral_queryset,  get_concept_total, get_full_catastral_queryset, calcular_igv_simple, obtener_calle, obtener_billing_type, get_morosos_queryset, to_none_if_empty, clean_value, to_none_if_empty_has_meter, to_decimal_or_none, generar_periodos, format_period, generate_daily_report, generar_codigo_medidor_unico, procesar_pago
from .core.mixins import TenantSafeMixin

import re

import io
import pandas as pd
import os
import zipfile
import uuid
import mercadopago

class CustomPagination(PageNumberPagination):

    page_size = 5  # Número de registros por página
    page_size_query_param = 'page_size'  # Permite cambiar el tamaño desde la URL
    max_page_size = 100  # Tamaño máximo permitido

class ServiceChargeViewSet(viewsets.ModelViewSet): 
    
    queryset = ServiceCharge.objects.all().order_by('-status','-id')
    serializer_class = ServiceChargeSerializer
    filter_backends = [DjangoFilterBackend,filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['customer', 'status']  

    def perform_destroy(self, instance):

        if instance.status == 'paid':
            raise ValidationError(
                {"error" :'No se puede eliminar un cargo que ya ha sido pagado.'}
            )

        instance.delete()

class ZonaViewSet(TenantSafeMixin,viewsets.ModelViewSet):

    queryset = Zona.objects.all().order_by('id')
    serializer_class = ZonaSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['codigo','name']

    @action(detail=False, methods=['post'])
    def importar_manzanas(self, request):

        archivo = request.FILES.get('file')

        if not archivo:
            return Response(
                {'error': 'Debe subir un archivo Excel'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:

            df = pd.read_excel(archivo)

            # convertir a numérico
            df['cr3_num'] = pd.to_numeric(df['cr3'], errors='coerce')
            df['cr4_num'] = pd.to_numeric(df['cr4'], errors='coerce')

            # ordenar primero por zona luego por manzana
            df = df.sort_values(
                by=['cr3_num', 'cr4_num'],
                ascending=[True, True]
            )

            total = 0

            for _, row in df.iterrows():

                zona_codigo = clean_value(row['cr3'])
                manzana_codigo = clean_value(row['cr4'])

                try:
                    zona = Zona.objects.get(codigo=zona_codigo)

                    _, created = Manzana.objects.get_or_create(
                        zona=zona,
                        codigo=manzana_codigo
                    )

                    if created:
                        total += 1

                except Zona.DoesNotExist:
                    continue

            return Response({
                'message': 'Importación completada',
                'manzanas_creadas': total
            })

        except Exception as e:

            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ManzanaViewSet(TenantSafeMixin,viewsets.ModelViewSet):

    queryset = Manzana.objects.all().order_by('id')
    serializer_class = ManzanaSerializer
    pagination_class = CustomPagination

    @action(detail=False, methods=['get'])
    def por_zona(self, request):

        zona_id = request.query_params.get('zona')

        queryset = Manzana.objects.filter(zona_id=zona_id).values('id', 'codigo')

        return Response(queryset)

class CustomerViewSet(TenantSafeMixin, GlobalPermissionMixin, viewsets.ModelViewSet):
  
    serializer_class = CustomerSerializer
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend,filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['codigo', 'full_name', 'number']
    filterset_fields = ['codigo','zona','calle','state','has_meter']  
    ordering_fields = ['total_debt','codigo']  # 👈 habilitamos orden
  
    def get_queryset(self):

        queryset = Customer.objects.filter(
            status=True
        ).annotate(
            total_debt=Coalesce(
                Sum(
                    'debts__amount',
                    filter=Q(
                        debts__paid=False,
                        debts__is_refinanced=False
                    )
                ),
                0,
                output_field=DecimalField(
                    max_digits=10,
                    decimal_places=2
                )
            )
        ).order_by('-codigo')

        # Solo para chilca
        if self.request.tenant.schema_name == 'chilca':

            queryset = queryset.annotate(

                mz_number=Cast(
                    'manzana__codigo',
                    IntegerField()
                ),

           

            ).order_by(
                'sector',
                'mz_number',
                # 'predio_number',
            )

        return queryset

    def create(self, request, *args, **kwargs):

        data = request.data
        tenant = request.tenant.schema_name

        code_number = 10
        next_code_fixed = "0000000001"

        if tenant == 'pangoa' or tenant == 'chilca' or tenant == 'demo':

           code_number = 5
           next_code_fixed = "00001"

        elif tenant == 'sanmarcos':

           code_number = 8
           next_code_fixed = "00000001"
        
        try:

            with transaction.atomic():
                # Obtener último código y sumar 1

                last_code = Customer.objects.aggregate(max_code=Max('codigo'))['max_code']

                if last_code:

                    next_code = str(int(last_code) + 1).zfill(code_number)

                else:

                    next_code = next_code_fixed

                # Asignar el nuevo código al cliente
                data['codigo'] = next_code

                customer_serializer = CustomerSerializer(data=data)
                customer_serializer.is_valid(raise_exception=True)
                customer = customer_serializer.save()

                return Response(CustomerSerializer(customer).data, status=status.HTTP_201_CREATED)

        except Exception as e:

            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
  
    def update(self, request, *args, **kwargs):

        self.required_action = "edit"
        self.check_global_permission(request)

        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):

        # Validar permiso global para eliminar
        self.required_action = "delete"
        self.check_global_permission(request)

        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def next_codigo(self, request):
        last = Customer.objects.aggregate(max_codigo=Max('codigo'))

        if last['max_codigo']:
            next_num = int(last['max_codigo']) + 1
        else:
            next_num = 1

        codigo = str(next_num).zfill(5)

        return Response({"codigo": codigo})

    @action(detail=False, methods=['get'])
    def supply_number(self, request):
        last = Customer.objects.aggregate(max_codigo=Max('supply_number'))

        if last['max_codigo']:
            next_num = int(last['max_codigo']) + 1
        else:
            next_num = 1

        codigo = str(next_num).zfill(5)

        return Response({"supply_number": codigo})

    @action(detail=False, methods=["get"], url_path="by-code")
    def by_code_and_dni(self, request):
        codigo = request.query_params.get("codigo")
        # dni = request.query_params.get("dni")
        #  or not dni
        if not codigo:
            return Response(
                {"error": "Debe proporcionar codigo de suministro"},
                status=status.HTTP_400_BAD_REQUEST
            )
        # , number=dni
        try:
            customer = Customer.objects.get(codigo=codigo)
        except Customer.DoesNotExist:
            return Response(
                {"error": "Cliente no encontrado"},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = CustomerWithDebtsSerializer(customer)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path='report/debt')
    def report(self,request):

        calle_id = request.query_params.get("calle")
        zona_id = request.query_params.get("zona")

        debts = Debt.objects.filter(paid=False)

        data = []

        total_general = Decimal("0.00")
        customers = Customer.objects.all()

        calle = None
        zona = None
        if calle_id:

            calle = Calle.objects.get(pk=calle_id)
            customers = customers.filter(calle_id=calle_id)

        if zona_id:

            zona = Zona.objects.get(pk=zona_id)
            customers = customers.filter(zona_id=zona_id)

        for customer in customers:

            customers_debts = debts.filter(customer=customer)

            if not customers_debts.exists():

               continue
               
            sumary = customers_debts.aggregate(
                total=Sum("amount"),
                min_period=Min("period"),
                max_period=Max("period")
            )

            total_general += sumary["total"] or Decimal("0.00")

            data.append({

                "customer" : customer,
                "min_period" : sumary["min_period"],
                "max_period" : sumary["max_period"],
                "total" : sumary["total"]

            })
     

        html_string = render_to_string("customer/report.html",{

            "data":data,
            "total_general":total_general,
            "date":datetime.now(),
            "calle": calle,
            "zona": zona

        })

        pdf = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'filename="reporte_global_deudas.pdf"'
        return response

    @action(detail=True, methods=['get'], url_path='report/debt-history')
    def report_debt_history(self, request, pk=None, **kwargs):

        customer = self.get_object()

        ####################################################
        # DEUDAS
        ####################################################

        debts = customer.debts.filter(paid=False, is_refinanced=False).order_by('period')

        debts_by_year = defaultdict(list)

        for debt in debts:

            debts_by_year[
                debt.period.year
            ].append(debt)

        ####################################################
        # REFINANCIAMIENTOS
        ####################################################

        refinancings = DebtRefinancing.objects.filter(
            customer=customer
        ).prefetch_related('installment_details')

        ####################################################
        # TOTALES DEUDAS
        ####################################################

        total_debt = (
            debts.aggregate(
                total=Sum('amount')
            )['total']
            or 0
        )

        total_paid = (
            debts.filter(
                paid=True
            ).aggregate(
                total=Sum('amount')
            )['total']
            or 0
        )

        total_pending = (
            total_debt - total_paid
        )

        ####################################################
        # TOTALES REFINANCIAMIENTO
        ####################################################

        total_refinanced = (
            refinancings.aggregate(
                total=Sum('total_amount_with_interest')
            )['total']
            or 0
        )

        total_refinancing_paid = (
            RefinancingInstallment.objects.filter(
                refinancing__customer=customer,
                paid=True
            ).aggregate(
                total=Sum('total_amount')
            )['total']
            or 0
        )

        total_refinancing_pending = (
            RefinancingInstallment.objects.filter(
                refinancing__customer=customer,
                paid=False
            ).aggregate(
                total=Sum('total_amount')
            )['total']
            or 0
        )

        total_full_paid = total_refinancing_paid + total_paid

        ####################################################
        # TEMPLATE
        ####################################################

        html_string = render_to_string(

            'customer/customer_debt_history.html',

            {

                'customer': customer,

                'debts_by_year': dict(
                    sorted(debts_by_year.items())
                ),

                'refinancings': refinancings,

                'total_debt': total_debt,

                'total_paid': total_paid,

                'total_pending': total_pending,

                'total_refinanced': total_refinanced,

                'total_refinancing_paid':
                    total_refinancing_paid,

                'total_refinancing_pending':
                    total_refinancing_pending,

                'today': datetime.now(),
                'total_full_paid' : total_full_paid

            }

        )

        pdf = HTML(
            string=html_string
        ).write_pdf()

        filename = (
            f"Historial_"
            f"{customer.full_name.replace(' ', '_')}.pdf"
        )

        response = HttpResponse(
            pdf,
            content_type='application/pdf'
        )

        response[
            'Content-Disposition'
        ] = f'inline; filename="{filename}"'

        return response

    @action(detail=False, methods=['get'])
    def export_template(self, request):

        year = int(request.GET.get("year", date.today().year))
        customers = get_full_catastral_queryset()
        wb = Workbook()
        ws = wb.active
        ws.title = f"PADRON GENERAL {year}"

        months = [
            "Enero",
            "Febrero",
            "Marzo",
            "Abril",
            "Mayo",
            "Junio",
            "Julio",
            "Agosto",
            "Septiembre",
            "Octubre",
            "Noviembre",
            "Diciembre",
        ]

        headers = [
            "Código",
            "Cliente",
            "Estado",
            "Observación",
            "Medidor",
            "Dirección",
            "CR1",
            "CR2",
            "CR3",
            "CR4",
            "CR5",
        ] + months

        # Encabezados
        for col_num, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)


        row = 2

        for customer in customers:

            ws.cell(row=row, column=1, value=customer.codigo)

            ws.cell(row=row, column=2, value=customer.full_name)

            ws.cell(
                row=row,
                column=3,
                value=customer.get_state_display()
            )

            ws.cell(
                row=row,
                column=4,
                value=customer.observation or ""
            )

            # 👇 ahora sale del annotate
            ws.cell(
                row=row,
                column=5,
                value=customer.meter_code or ""
            )

            ws.cell(
                row=row,
                column=6,
                value=customer.address
            )

            # Catastro
            ws.cell(row=row, column=7, value=customer.provincia)
            ws.cell(row=row, column=8, value=customer.distrito)
            ws.cell(row=row, column=9, value=customer.sector)
            ws.cell(row=row, column=10, value=customer.manzana.codigo if customer.manzana else '')
            ws.cell(row=row, column=11, value=customer.predio)

            readings = customer.readings.filter(
                period__year=year
            )

            readings_map = {
                r.period.month: r.current_reading
                for r in readings
            }

            # meses empiezan en columna 12
            for month in range(1, 13):

                value = readings_map.get(month, "")

                ws.cell(
                    row=row,
                    column=month + 11,
                    value=float(value) if value else ""
                )

            row += 1

        # Ajustar tamaño columnas
        for column_cells in ws.columns:

            length = max(len(str(cell.value or "")) for cell in column_cells)

            ws.column_dimensions[
                column_cells[0].column_letter
            ].width = length + 5

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response[
            'Content-Disposition'
        ] = f'attachment; filename="lecturas_{year}.xlsx"'

        wb.save(response)

        return response

    @action(detail=True, methods=['get'],url_path='refinancings')
    def refinancings(self, request, pk=None):

        customer = self.get_object()

        queryset = DebtRefinancing.objects.filter(
            customer=customer
        ).prefetch_related(

            'details__debt',

            Prefetch(
                'installment_details',
                queryset=RefinancingInstallment.objects.order_by('number')
            )

        ).order_by('-created_at')

        serializer = DebtRefinancingSerializer(
            queryset,
            many=True
        )

        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def import_excel_readings(self, request):

        file = request.FILES.get('file')

        if not file:
            return Response({
                'error': 'No se proporcionó un archivo.'
            }, status=400)

        # ==========================================
        # LEER EXCEL
        # ==========================================
        try:

            df = pd.read_excel(
                file,
                sheet_name=2,
                engine='openpyxl'
            )

        except Exception as e:

            return Response({
                'error': f'Error al leer el archivo: {str(e)}'
            }, status=400)

        # ==========================================
        # LIMPIAR COLUMNAS
        # ==========================================
        df.columns = df.columns.str.strip().str.upper()

        created = 0
        skipped = 0

        # ==========================================
        # RECORRER FILAS
        # ==========================================
        for _, row in df.iterrows():

            # ==========================================
            # CODIGO CLIENTE
            # ==========================================
            codigo = str(
                clean_value(row.get('SUMINISTRO'))
            ).strip()

            observation = clean_value(
                row.get('OBSERVACIONES')
            )

            observation = observation.strip() if observation else None
        
            if not codigo:
                skipped += 1
                continue

            # ==========================================
            # CLIENTE
            # ==========================================
            customer = Customer.objects.filter(
                codigo=codigo,
                state='active'
            ).first()

            if not customer:
                skipped += 1
                continue

            customer.observation = observation
            customer.save()

            # ==========================================
            # MEDIDOR ACTIVO
            # ==========================================
            assignment = MeterAssignment.objects.filter(
                customer=customer
            ).first()

            if not assignment or not assignment.meter:
                skipped += 1
                continue

            meter = assignment.meter

            # ==========================================
            # LECTURA DE MAYO
            # ==========================================
            current_reading = to_decimal_or_none(
                row.get('LEC. MAYO')
            )

            if current_reading is None:
                skipped += 1
                continue

            # ==========================================
            # PERIODO
            # ==========================================
            period_date = date(2026, 5, 1)

            # ==========================================
            # VALIDAR SI YA EXISTE
            # ==========================================
            exists = Reading.objects.filter(
                customer=customer,
                period=period_date
            ).exists()

            if exists:

                print(
                    f"YA EXISTE -> {customer.codigo}"
                )

                skipped += 1
                continue

            # ==========================================
            # OBTENER ULTIMA LECTURA
            # ==========================================
            last_reading = Reading.objects.filter(
                customer=customer
            ).order_by('-period').first()

            # ==========================================
            # CALCULAR CONSUMO
            # ==========================================
            if last_reading:

                prev_value = last_reading.current_reading

                consumption = (
                    current_reading - prev_value
                )

                # evitar negativos
                if consumption < 0:
                    consumption = Decimal('0.000')

            else:

                # primera lectura histórica
                prev_value = current_reading
                consumption = Decimal('0.000')

   
            try:

                with transaction.atomic():

                    # ==========================================
                    # CREAR LECTURA
                    # ==========================================
                    reading = Reading.objects.create(
                        customer=customer,
                        meter=meter,

                        period=period_date,

                        current_reading=current_reading,
                        previous_reading=prev_value,
                        consumption=consumption,

                        status='normal',
                        has_meter=True
                    )

                    created += 1

                    print(
                        f"CREADO -> "
                        f"{customer.codigo} | "
                        f"Anterior: {prev_value} | "
                        f"Actual: {current_reading} | "
                        f"Consumo: {consumption}"
                    )

            except IntegrityError:

                print(
                    f"DUPLICADO -> "
                    f"{customer.codigo}"
                )

                skipped += 1

            except Exception as e:

                print(
                    f"ERROR -> "
                    f"{customer.codigo} -> {str(e)}"
                )

                skipped += 1

        return Response({
            "message": "Importación completada",
            "readings_created": created,
            "skipped": skipped
        }, status=status.HTTP_201_CREATED)

class CashBoxViewSet(TenantSafeMixin,viewsets.ModelViewSet):
    
    queryset = CashBox.objects.all()
    serializer_class = CashBoxSerializer

    @action(detail=True, methods=["post"], url_path='close-cash')
    def confirm_daily_report(self, request, pk=None):
        cashbox = self.get_object()
        date = request.data.get("date") or str(localdate())

        report = generate_daily_report(cashbox, date)
        report.confirmed = True
        report.save()

        return Response({"message": f"Caja del {report.date} confirmada", "closing_balance": report.closing_balance})

    @action(detail=True, methods=["get"])
    def report(self, request, pk=None, **kwargs):

        cashbox = self.get_object()
        user = User.objects.filter(id=cashbox.user_id).first()
        # 📌 Filtros de fechas
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        date_param = request.query_params.get("date")

        movimientos = cashbox.movements.all()
        egresos = cashbox.outflows.all()

        # SYSTEM_KEYS = (
        #     ("price_water", "Precio agua"),
        #     ("price_sewer", "Precio alcantarillado"),
        #     ("price_fixed_charge", "Precio cargo fijo"),
        #     ("price_clean", "Precio limpieza"),
        #     ("price_igv", "Precio Igv"),
        # )

        PERIOD_CONCEPTS = {
            "price_water",
            "price_sewer",
            "price_fixed_charge",
        }

        if start_date and end_date:

            try:
                start = datetime.strptime(start_date, "%Y-%m-%d").date()
                end = datetime.strptime(end_date, "%Y-%m-%d").date()
            except ValueError:
                return Response({"error": "Formato de fecha inválido (use YYYY-MM-DD)"}, status=400)

            movimientos = movimientos.filter(created_at__date__range=(start, end))
            egresos = egresos.filter(created_at__date__range=(start, end))
            reporte_tipo = f"Reporte entre {start} y {end}"

        else:

            if date_param:
                try:
                    fecha = datetime.strptime(date_param, "%Y-%m-%d").date()
                except ValueError:
                    return Response({"error": "Formato de fecha inválido (use YYYY-MM-DD)"}, status=400)
            else:
                fecha = localdate()

            movimientos = movimientos.filter(created_at__date=fecha)
            egresos = egresos.filter(created_at__date=fecha)
            reporte_tipo = f"Reporte diario - {fecha}"

        # ==========================
        # AGRUPACIÓN POR CONCEPTO
        # ==========================
        conceptos_dict = defaultdict(list)

        for mov in movimientos.select_related("concept", "invoice_payment__invoice__customer"):
            concepto = mov.concept.name
            conceptos_dict[concepto].append(mov)

        conceptos_data = []

        for concepto, movs in conceptos_dict.items():
            facturas_dict = {}

            for m in movs:
                inv = m.invoice_payment.invoice if m.invoice_payment else None

                if not inv or inv.status == "cancelled":
                    continue  # ignoramos facturas anuladas

                key = inv.id
                if key not in facturas_dict:
                    facturas_dict[key] = {
                        "code": inv.code,
                        "number_reference" : inv.number_reference,
                        "date": inv.date,
                        "cliente": inv.customer.full_name,
                        "cliente_codigo": inv.customer.codigo,
                        "direccion": inv.customer.address,
                        "pagos": defaultdict(float),
                        "total": 0,
                        "periodo": "",  # 👈 string vacío por defecto
                    }

                    # 📌 Solo calcular periodo si el concepto es 001, 002 o 003
                    if m.concept.system_key in PERIOD_CONCEPTS:

                        periodos = list(
                            inv.invoice_debts.select_related("debt")
                            .values_list("debt__period", flat=True)
                        )

                        if periodos:
                            periodos = sorted(periodos)
                            if len(periodos) == 1:
                                facturas_dict[key]["periodo"] = format_period(periodos[0])
                            else:
                                facturas_dict[key]["periodo"] = (
                                    f"{format_period(periodos[0])} - {format_period(periodos[-1])}"
                                )

                metodo = m.invoice_payment.method if m.invoice_payment else "Desconocido"
                facturas_dict[key]["pagos"][metodo] += float(m.total)
                facturas_dict[key]["total"] += float(m.total)

            # Convertir a lista
            facturas_list = []
            total_concepto = 0
            for f in facturas_dict.values():
                total_concepto += f["total"]
                f["pagos"] = dict(f["pagos"])  # pasar defaultdict a dict normal
                facturas_list.append(f)

            conceptos_data.append({
                "concepto": concepto,
                "total": total_concepto,
                "facturas": facturas_list,
            })

        total_general = sum(c["total"] for c in conceptos_data)

        # Agrupar por método de pago
        metodo_data = []

        payments = cashbox.payments.select_related("invoice")

        if start_date and end_date:

            payments = payments.filter(
                created_at__date__range=(start, end)
            )

        else:

            payments = payments.filter(
                created_at__date=fecha
            )

        metodo_dict = defaultdict(list)

        for pay in payments:

            if pay.invoice.status == "cancelled":
                continue

            metodo_dict[pay.method].append(pay)

        for metodo, pays in metodo_dict.items():

            total_metodo = sum([p.total for p in pays])

            metodo_data.append({
                "metodo": dict(InvoicePayment.PAYMENT_METHODS).get(metodo, metodo),
                "total": total_metodo,
                "movimientos": pays
            })

        total_metodos = sum([
            m["total"] for m in metodo_data
        ])

        html_string = render_to_string("reports/caja/daily.html", {
            "cashbox": cashbox,
            "user" : user,
            "conceptos": conceptos_data,
            "total_general": total_general,
            "total_metodos": total_metodos,
            "reporte_tipo": reporte_tipo,
            "metodos": metodo_data,
            "report": cashbox,
        })

        pdf = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'filename="reporte_caja_{cashbox.id}.pdf"'
        return response

class DailyCashReportViewSet(TenantSafeMixin,viewsets.ModelViewSet):
    
    queryset = DailyCashReport.objects.all().order_by('-date')
    serializer_class = DailyCashReportSerializer

    @action(detail=True, methods=["get"])
    def report(self, request, pk=None, **kwargs):

        daily_cash = self.get_object()

        cashbox = daily_cash.cashbox
        user = User.objects.filter(id=cashbox.user_id).first()
        fecha = daily_cash.date

        movimientos = cashbox.movements.filter(created_at__date=fecha)
        reporte_tipo = f"Reporte - {fecha}"

        # ==========================
        # AGRUPACIÓN POR CONCEPTO
        # ==========================
        conceptos_dict = defaultdict(list)

        PERIOD_CONCEPTS = {
            "price_water",
            "price_sewer",
            "price_fixed_charge",
        }

        for mov in movimientos.select_related("concept", "invoice_payment__invoice__customer"):
            concepto = mov.concept.name
            conceptos_dict[concepto].append(mov)

        conceptos_data = []

        for concepto, movs in conceptos_dict.items():
            facturas_dict = {}

            for m in movs:
                inv = m.invoice_payment.invoice if m.invoice_payment else None

                if not inv or inv.status == "cancelled":
                    continue  # ignoramos facturas anuladas

                key = inv.id
                if key not in facturas_dict:
                    facturas_dict[key] = {
                        "code": inv.code,
                        "number_reference" : inv.number_reference,
                        "date": inv.date,
                        "cliente": inv.customer.full_name,
                        "cliente_codigo": inv.customer.codigo,
                        "direccion": inv.customer.address,
                        "pagos": defaultdict(float),
                        "total": 0,
                        "periodo": "",  # 👈 string vacío por defecto
                    }

                    # 📌 Solo calcular periodo si el concepto es 001, 002 o 003
                    if m.concept.system_key in PERIOD_CONCEPTS:
                        periodos = list(
                            inv.invoice_debts.select_related("debt")
                            .values_list("debt__period", flat=True)
                        )

                        if periodos:
                            periodos = sorted(periodos)
                            if len(periodos) == 1:
                                facturas_dict[key]["periodo"] = format_period(periodos[0])
                            else:
                                facturas_dict[key]["periodo"] = (
                                    f"{format_period(periodos[0])} - {format_period(periodos[-1])}"
                                )

                metodo = m.invoice_payment.method if m.invoice_payment else "Desconocido"
                facturas_dict[key]["pagos"][metodo] += float(m.total)
                facturas_dict[key]["total"] += float(m.total)

            # Convertir a lista
            facturas_list = []
            total_concepto = 0
            for f in facturas_dict.values():
                total_concepto += f["total"]
                f["pagos"] = dict(f["pagos"])  # pasar defaultdict a dict normal
                facturas_list.append(f)

            conceptos_data.append({
                "concepto": concepto,
                "total": total_concepto,
                "facturas": facturas_list,
            })

        total_general = sum(c["total"] for c in conceptos_data)

        # Agrupar por método de pago
        metodo_data = []

        payments = cashbox.payments.select_related("invoice")

        payments = payments.filter(
            created_at__date=fecha
        )

        metodo_dict = defaultdict(list)

        for pay in payments:

            if pay.invoice.status == "cancelled":
                continue

            metodo_dict[pay.method].append(pay)

        for metodo, pays in metodo_dict.items():

            total_metodo = sum([p.total for p in pays])

            metodo_data.append({
                "metodo": dict(InvoicePayment.PAYMENT_METHODS).get(metodo, metodo),
                "total": total_metodo,
                "movimientos": pays
            })

        total_metodos = sum([
            m["total"] for m in metodo_data
        ])

        html_string = render_to_string("reports/caja/daily.html", {
            "cashbox": cashbox,
            "user": user,
            "conceptos": conceptos_data,
            "total_general": total_general,
            "total_metodos": total_metodos,
            "reporte_tipo": reporte_tipo,
            "metodos": metodo_data,
            "report": daily_cash,
        })

        pdf = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'filename="reporte_caja_{cashbox.id}.pdf"'
        return response

class WaterMeterViewSet(TenantSafeMixin,viewsets.ModelViewSet):
    
    queryset = WaterMeter.objects.all().order_by('-id')
    serializer_class = WaterMeterSerializer
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend,filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['code']
    filterset_fields = ['status']  

    @action(detail=True, methods=['post'])
    def assign(self, request, pk=None):

        meter = self.get_object()

        if meter.status != 'available':
            return Response({"error": "No disponible"}, status=400)

        customer_id = request.data.get('customer')
        installation_date = request.data.get('installation_date')

        MeterAssignment.objects.create(
            meter=meter,
            customer_id=customer_id,
            installation_date=installation_date
        )

        meter.status = 'installed'
        meter.save()

        serializer = WaterMeterSerializer(meter)

        return Response(
            serializer.data,
            status=status.HTTP_201_CREATED
        )

    @action(detail=True, methods=['post'])
    def remove(self, request, pk=None):

        meter = self.get_object()

        if meter.status != 'installed':
            return Response({"error": "El medidor no está instalado"}, status=400)

        assignment = meter.assignments.filter(is_active=True).first()

        if not assignment:
            return Response({"error": "No hay asignación activa"}, status=400)

        assignment.is_active = False
        assignment.removal_date = request.data.get('removal_date') or date.today()
        assignment.save()

        meter.status = 'removed'
        meter.save()

        return Response({"ok": True})

    @action(detail=True, methods=['post'])
    def mark_damaged(self, request, pk=None):

        meter = self.get_object()

        if meter.status != 'installed':
            return Response({"error": "Solo medidores instalados pueden dañarse"}, status=400)

        assignment = meter.assignments.filter(is_active=True).first()

        if assignment:
            assignment.is_active = False
            assignment.removal_date = date.today()
            assignment.save()

        meter.status = 'damaged'
        meter.save()

        return Response({"ok": True})

    @action(detail=True, methods=['post'])
    def send_to_maintenance(self, request, pk=None):

        meter = self.get_object()

        if meter.status != 'damaged':
            return Response({"error": "Solo medidores dañados"}, status=400)

        meter.status = 'maintenance'
        meter.save()

        return Response({"ok": True})

    @action(detail=True, methods=['post'])
    def release(self, request, pk=None):

        meter = self.get_object()

        if meter.status != 'maintenance':
            return Response({"error": "No está en mantenimiento"}, status=400)

        meter.status = 'available'
        meter.save()

        return Response({"ok": True})

    @action(detail=True, methods=['post'])
    def return_to_stock(self, request, pk=None):

        meter = self.get_object()

        if meter.status != 'removed':
            return Response({"error": "Solo medidores retirados"}, status=400)

        meter.status = 'available'
        meter.save()

        return Response({"ok": True})

    @action(detail=False, methods=['post'])
    def import_excel(self, request):

        file = request.FILES.get('file')

        if not file:
            return Response(
                {"error": "No se envió ningún archivo."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:

            # Leer Excel
            df = pd.read_excel(file)

            # Validar columna
            if 'codmedidor' not in df.columns:
                return Response(
                    {"error": "La columna 'codmedidor' no existe."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            created = 0
            repeated = []

            for _, row in df.iterrows():

                value = row.get('codmedidor')
                suministro = row.get('SUMINISTRO')

                if pd.isna(value):
                    continue

                code = str(value).strip()

                # ignorar valores
                if code.upper() in ['DIRECTO', 'DESAGUE', 'SIN MEDIDO']:
                    continue

                original_code = code

                # si existe, generar -R1, -R2, etc.
                if WaterMeter.objects.filter(code=code).exists():

                    counter = 1

                    while WaterMeter.objects.filter(
                        code=f"{original_code}-R{counter}"
                    ).exists():
                        counter += 1

                    code = f"{original_code}-R{counter}"

                    repeated.append({
                        "suministro": suministro,
                        "original": original_code,
                        "new_code": code
                    })

                WaterMeter.objects.create(
                    code=code,
                    status='available'
                )

                created += 1

            return Response({
                "message": "Importación completada.",
                "created": created,
                "repeated": repeated
            })

        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def available(self, request):

        meters = WaterMeter.objects.filter(
            status='available'
        ).order_by('code')

        serializer = self.get_serializer(meters, many=True)

        return Response(serializer.data)

class MeterAssignmentViewSet(TenantSafeMixin, viewsets.ModelViewSet):

    serializer_class = MeterAssignmentSerializer
    pagination_class = CustomPagination

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter
    ]

    search_fields = ['meter__code']
    
    filterset_fields = [
        'customer__codigo',
        'customer__state',
        'customer__zona',
        'customer__manzana'
    ]

    def get_queryset(self):

        month = self.request.query_params.get('month')

        if not month:
            return MeterAssignment.objects.none()

        year, month = map(int, month.split('-'))

        period_date = date(year, month, 1)

        queryset = get_catastral_queryset(period_date)

        # # SOLO MEDIDORES SIN LECTURA
        # if pending == 'si':
        #     queryset = queryset.filter(
        #         has_current_reading=False
        #     )

        return queryset

    @action(detail=False, methods=['get'])
    def export_template(self, request):

        month = request.query_params.get('month')

        if not month:
            return HttpResponse(
                'Debe enviar el parámetro month',
                status=400
            )

        year, month_number = map(int, month.split('-'))

        period_date = date(year, month_number, 1)

        assignments = get_catastral_queryset(period_date)

        wb = Workbook()

        ws = wb.active

        ws.title = f"LECTURAS {month}"

        months = [
            "Enero",
            "Febrero",
            "Marzo",
            "Abril",
            "Mayo",
            "Junio",
            "Julio",
            "Agosto",
            "Septiembre",
            "Octubre",
            "Noviembre",
            "Diciembre",
        ]

        current_month_name = months[month_number - 1]

        headers = [

            "Código",
            "Cliente",
            "Estado",
            "Observación",
            "Medidor",
            "Dirección",

            # Catastro
            "CR1",
            "CR2",
            "CR3",
            "CR4",
            "CR5",

            # Lectura anterior
            "Última lectura",
            "Último periodo",

            # Lectura actual
            f"Lectura {current_month_name} {year}",

        ]

        # =========================
        # ESTILOS
        # =========================

        green_fill = PatternFill(
            start_color="C6EFCE",
            end_color="C6EFCE",
            fill_type="solid"
        )

        yellow_fill = PatternFill(
            start_color="FFF3CD",
            end_color="FFF3CD",
            fill_type="solid"
        )

        # =========================
        # HEADERS
        # =========================

        for col_num, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col_num)

            cell.value = header

            cell.font = Font(bold=True)

        # =========================
        # DATA
        # =========================

        row = 2

        for assignment in assignments:

            customer = assignment.customer

            meter = assignment.meter

            # Código
            ws.cell(
                row=row,
                column=1,
                value=customer.codigo
            )

            # Cliente
            ws.cell(
                row=row,
                column=2,
                value=customer.full_name
            )

            # Estado
            ws.cell(
                row=row,
                column=3,
                value=customer.get_state_display()
            )

            # Observación
            ws.cell(
                row=row,
                column=4,
                value=customer.observation or ""
            )

            # Medidor
            ws.cell(
                row=row,
                column=5,
                value=meter.code if meter else ""
            )

            # Dirección
            ws.cell(
                row=row,
                column=6,
                value=customer.address
            )

            # =========================
            # CATASTRO
            # =========================

            ws.cell(
                row=row,
                column=7,
                value=customer.provincia
            )

            ws.cell(
                row=row,
                column=8,
                value=customer.distrito
            )

            ws.cell(
                row=row,
                column=9,
                value=customer.sector
            )

            ws.cell(
                row=row,
                column=10,
                value=(
                    customer.manzana.codigo
                    if customer.manzana
                    else customer.mz or ""
                )
            )

            ws.cell(
                row=row,
                column=11,
                value=customer.predio
            )

            # =========================
            # LECTURA ANTERIOR
            # =========================

            ws.cell(
                row=row,
                column=12,
                value=(
                    float(assignment.previous_reading)
                    if assignment.previous_reading is not None
                    else ""
                )
            )

            ws.cell(
                row=row,
                column=13,
                value=(
                    assignment.previous_period.strftime('%Y-%m')
                    if assignment.previous_period
                    else ""
                )
            )

            # =========================
            # LECTURA ACTUAL
            # =========================

            current_cell = ws.cell(
                row=row,
                column=14,
                value=(
                    float(assignment.current_reading_value)
                    if getattr(
                        assignment,
                        'current_reading_value',
                        None
                    ) is not None
                    else ""
                )
            )

            # Ya registrado
            if assignment.has_current_reading:

                current_cell.fill = green_fill

            # Sin lectura anterior
            elif not assignment.previous_reading:

                current_cell.fill = yellow_fill

            row += 1

        # =========================
        # AUTO WIDTH
        # =========================

        for column_cells in ws.columns:

            length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            ws.column_dimensions[
                column_cells[0].column_letter
            ].width = length + 5

        # =========================
        # RESPONSE
        # =========================

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response[
            'Content-Disposition'
        ] = (
            f'attachment; '
            f'filename="lecturas_{month}.xlsx"'
        )

        wb.save(response)

        return response

    @action(detail=False, methods=['get'])
    def export_history(self, request):

        year = request.query_params.get( 'year',datetime.now().year)

        if not year:
            return HttpResponse(
                'Debe enviar el parámetro year',
                status=400
            )

        year = int(year)

        assignments = (
            MeterAssignment.objects
            .select_related(
                'customer',
                'meter'
            )
            .filter(
                customer__state='active'
            )
            .order_by(
                'customer__sector',
                'customer__manzana__codigo',
                'customer__predio'
            )
        )

        wb = Workbook()

        ws = wb.active

        ws.title = f"HISTORIAL {year}"

        months = [
            "Enero",
            "Febrero",
            "Marzo",
            "Abril",
            "Mayo",
            "Junio",
            "Julio",
            "Agosto",
            "Septiembre",
            "Octubre",
            "Noviembre",
            "Diciembre",
        ]

        headers = [

            "Código",
            "Cliente",
            "Estado",
            "Observación",
            "Medidor",
            "Dirección",

            # Catastro
            "CR1",
            "CR2",
            "CR3",
            "CR4",
            "CR5",

        ] + months

        # =========================
        # HEADERS
        # =========================

        for col_num, header in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col_num)

            cell.value = header

            cell.font = Font(bold=True)

        # =========================
        # DATA
        # =========================

        row = 2

        for assignment in assignments:

            customer = assignment.customer

            meter = assignment.meter

            # Código
            ws.cell(
                row=row,
                column=1,
                value=customer.codigo
            )

            # Cliente
            ws.cell(
                row=row,
                column=2,
                value=customer.full_name
            )

            # Estado
            ws.cell(
                row=row,
                column=3,
                value=customer.get_state_display()
            )

            # Observación
            ws.cell(
                row=row,
                column=4,
                value=customer.observation or ""
            )

            # Medidor
            ws.cell(
                row=row,
                column=5,
                value=meter.code if meter else ""
            )

            # Dirección
            ws.cell(
                row=row,
                column=6,
                value=customer.address
            )

            # =========================
            # CATASTRO
            # =========================

            ws.cell(
                row=row,
                column=7,
                value=customer.provincia
            )

            ws.cell(
                row=row,
                column=8,
                value=customer.distrito
            )

            ws.cell(
                row=row,
                column=9,
                value=customer.sector
            )

            ws.cell(
                row=row,
                column=10,
                value=(
                    customer.manzana.codigo
                    if customer.manzana
                    else customer.mz or ""
                )
            )

            ws.cell(
                row=row,
                column=11,
                value=customer.predio
            )

            # =========================
            # LECTURAS DEL AÑO
            # =========================

            readings = customer.readings.filter(
                period__year=year
            )

            readings_map = {
                r.period.month: {
                    "reading": r.current_reading,
                    "observation": r.observation or ""
                }
                for r in readings
            }

            # Enero -> Diciembre
            for month_number in range(1, 13):

                data = readings_map.get(month_number)

                if data:

                    reading = data["reading"]
                    observation = data["observation"]

                    cell_value = f"{float(reading)}"

                    if observation:
                        cell_value += f" ({observation})"

                else:
                    cell_value = ""

                ws.cell(
                    row=row,
                    column=month_number + 11,
                    value=cell_value
                )

            row += 1

        # =========================
        # AUTO WIDTH
        # =========================

        for column_cells in ws.columns:

            length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            ws.column_dimensions[
                column_cells[0].column_letter
            ].width = length + 5

        # =========================
        # RESPONSE
        # =========================

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response[
            'Content-Disposition'
        ] = (
            f'attachment; '
            f'filename="historial_lecturas_{year}.xlsx"'
        )

        wb.save(response)

        return response

class ReadingViewSet(TenantSafeMixin,viewsets.ModelViewSet):

    queryset = Reading.objects.all().order_by('period')
    serializer_class = ReadingSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = ReadingFilter

    def perform_update(self, serializer):
        instance = self.get_object()

        # 🔒 Bloqueo extra desde ViewSet
        if instance.paid:
            raise ValidationError(
                f"No se puede modificar la lectura de {instance.period.strftime('%Y-%m')} porque ya esta pagada."
            )

        return serializer.save()
    
    def perform_destroy(self, instance):

        # 🔒 No borrar si la deuda ya fue pagada
        if hasattr(instance, "debt") and instance.debt and instance.debt.paid:
            raise ValidationError({
                "error": "No se puede eliminar una lectura que ya está pagada."
            })

        # 🔒 No borrar si existen lecturas posteriores pagadas
        has_paid_next = Reading.objects.filter(
            customer=instance.customer,
            period__gt=instance.period,
            debt__paid=True
        ).exists()

        if has_paid_next:
            raise ValidationError({
                "error": "No se puede eliminar porque existen lecturas posteriores ya pagadas."
            })

        # Eliminar deuda asociada
        if hasattr(instance, "debt") and instance.debt:
            instance.debt.delete()

        # Eliminar lectura
        instance.delete()

        # # Obtener lectura previa
        # prev_reading = Reading.objects.filter(
        #     customer=customer,
        #     period__lt=period
        # ).order_by("-period").first()

        # prev_value = (
        #     prev_reading.current_reading
        #     if prev_reading
        #     else 0
        # )

        # # 🔄 Recalcular posteriores
        # for r in next_readings:

        #     r.previous_reading = prev_value
        #     r.consumption = r.current_reading - prev_value

        #     # Recalcular deuda
        #     if hasattr(r, "debt") and r.debt:

        #         r.debt.amount = (
        #             r.consumption *
        #             r.customer.category.price_water
        #         )

        #         r.debt.save()

        #     r.save()

        #     prev_value = r.current_reading

    @action(detail=False, methods=['get'], url_path='has-history/(?P<customer_id>[^/.]+)')
    def has_history(self, request, customer_id=None):

        period = request.GET.get('period')
        reading_id = request.GET.get('reading_id')

        readings = Reading.objects.filter(
            customer_id=customer_id
        )

        # =====================================
        # ÚLTIMA LECTURA
        # =====================================
        last_reading = readings.order_by('-period').first()

        # =====================================
        # VALIDAR LECTURAS POSTERIORES
        # =====================================
        has_next_readings = False
        has_paid_next_readings = False

        if period:

            next_readings = readings.filter(
                period__gt=period
            )

            # excluir lectura actual en edición
            if reading_id:

                next_readings = next_readings.exclude(
                    id=reading_id
                )

            has_next_readings = next_readings.exists()

            has_paid_next_readings = next_readings.filter(
                paid=True
            ).exists()

        # =====================================
        # SIN HISTORIAL
        # =====================================
        if not last_reading:

            return Response({

                'hasHistory': False,

                'hasNextReadings': False,

                'hasPaidNextReadings': False,

                'lastReading': None

            })

        # =====================================
        # RESPUESTA
        # =====================================
        return Response({

            'hasHistory': True,

            'hasNextReadings': has_next_readings,

            'hasPaidNextReadings': has_paid_next_readings,

            'lastReading': {

                'id': last_reading.id,

                'period': last_reading.period,

                'previous_reading': last_reading.previous_reading,

                'current_reading': last_reading.current_reading,

                'consumption': last_reading.consumption,

                'paid': last_reading.paid,

            }

        })

    @action(detail=True, methods=['get'])
    def receipt(self, request, pk=None, **kwargs):

        tenant = request.tenant.schema_name
        company = Company.objects.first()

        # reading_generation = (
        #     ReadingGeneration.objects
        #     .order_by('-period')
        #     .first()
        # )

        # print(reading_generation)

        # if reading_generation:

        #     current_period = reading_generation.period

        # else:
        #     # buscar el período anterior registrado
        #     previous_generation = (
        #         ReadingGeneration.objects
        #         .exclude(period=None)
        #         .order_by('-created_at')
        #         .first()
        #     )

        #     if not previous_generation:
        #         return Response(
        #             {"error": "No existe una generación de lecturas"},
        #             status=404
        #         )

        #     current_period = previous_generation.period

        # print(current_period)

        last_reading = Reading.objects.filter(customer_id=pk).first()
        current_period = last_reading.period
        reading = Reading.objects.filter(
            customer_id=pk,
            period=current_period
        ).first()

        debt = Debt.objects.filter(
            customer_id=pk,
            period=current_period
        ).first()

        if not reading:

            return Response({"error": "No hay lecturas registradas"}, status=404)
        
        # conceptos que se deben mostrar en el recibo
        master_concepts = CashConcept.objects.filter(is_master_view=True,state=True).order_by("id")

        # convertir detalles de deuda a diccionario
        detail_map = {
            detail.concept_id: detail.amount
            for detail in debt.details.all()
        }

        receipt_details = []

        for concept in master_concepts:

            receipt_details.append({
                "concept": concept,
                "amount": detail_map.get(concept.id, 0)
            })

        # Ruta al logo según el RUC
        logo_path = None

        if company and company.ruc:
           logo_path = request.build_absolute_uri(f"/media/{company.ruc}.jpeg")

        ###################################################### 
        #               DEUDAS ANTERIORES       
        ######################################################

        # obtener deudas anteriores no pagadas
        previous_debts = Debt.objects.filter(
            customer=reading.customer,
            paid=False,
            period__lt=reading.period,
            is_refinanced = False
        ).order_by("period")

        # Agrupar por año
        yearly_data = defaultdict(lambda: {"total": 0, "months": []})
        for d in previous_debts:
            year = d.period.year
            month = d.period.month
            yearly_data[year]["total"] += float(d.amount)
            yearly_data[year]["months"].append(month)

        grouped_debts = []
        for year, data in yearly_data.items():
            min_month = min(data["months"])
            max_month = max(data["months"])
            grouped_debts.append({
                "year": year,
                "total": f"{data['total']:.2f}",
                "from_month": format_date(date(year, min_month, 1), "MMMM", locale="es").capitalize(),
                "to_month": format_date(date(year, max_month, 1), "MMMM", locale="es").capitalize(),
            })

        grouped_debts.sort(key=lambda x: x["year"], reverse=True)
        total_previous_debt = previous_debts.aggregate(total=Sum("amount"))["total"] or 0


        ######################################################
        #           OTROS CONCEPTOS PENDIENTES
        ######################################################

        pending_services = ServiceCharge.objects.filter(
            customer=reading.customer,
            status='pending'
        ).exclude(
            invoice__status='cancelled'
        ).select_related("concept")

        total_other_services = pending_services.aggregate(
            total=Sum("amount")
        )["total"] or 0

        other_services = []

        for service in pending_services:

            other_services.append({
                "concept": service.concept.name,
                "description": service.description,
                "amount": service.amount,
                "created_at": service.created_at,
            })


        ######################################################
        #           CUOTA PENDIENTE REFINANCIAMIENTO
        ######################################################

        pending_installment = (
            RefinancingInstallment.objects
            .filter(
                refinancing__customer=reading.customer,
                paid=False,
                refinancing__paid=False
            )
            .select_related("refinancing")
            .order_by("due_date", "number")
            .first()
        )

        refinancing_data = None
        total_refinancing = 0

        if pending_installment:

            refinancing_data = {
                "number": pending_installment.number,
                "total_amount": pending_installment.total_amount,
                "due_date": pending_installment.due_date,
                "refinancing_id": pending_installment.refinancing.id,
            }

            total_refinancing = pending_installment.total_amount

        if debt.paid:

            total_general = (total_previous_debt + total_other_services + total_refinancing)

        else:

            total_general = ( debt.amount + total_previous_debt + total_other_services + total_refinancing)

        total_previous_debt_global = (total_previous_debt +  total_other_services + total_refinancing)

        # armamos la misma estructura que en el masivo
        readings_context = [{
            "debt": debt,
            "details": receipt_details,
            "reading": reading,
            "grouped_debts": grouped_debts,
            "total_previous_debt": total_previous_debt,
            "total_other_services": total_other_services,
            "total_previous_debt_global" : total_previous_debt_global,
            "other_services": other_services,
            "refinancing_data": refinancing_data,
            "total_general": total_general,
        }]
        

        if tenant == 'chilca':

            background_image = request.build_absolute_uri(f"/media/chilca.png")

            html = render_to_string("agua/chilca.html", {
                "readings_context": readings_context,
                "company": company,
                "company_logo": logo_path,
                "background_image" : background_image
            })

        else:

            html = render_to_string("agua/recibo.html", {
                "readings_context": readings_context,
                "company": company,
                "company_logo": logo_path,
            })

        pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename=recibo_{reading.customer.codigo}_{reading.period.strftime("%Y-%m")}.pdf"'
        return response

    @action(detail=False, methods=['get'])
    def get_progress(self, request):

        period = request.query_params.get("month")

        if not period:
            return Response({"error": "month requerido"}, status=400)

        year, month = map(int, period.split('-'))
        period_date = date(year, month, 1)

        readings = Reading.objects.filter(
            customer=OuterRef('pk'),
            period=period_date
        )

        qs = Customer.objects.filter(
            state='active',
            status=True
        ).annotate(
            has_reading=Exists(readings)
        )

        total = qs.count()
        registrados = qs.filter(has_reading=True).count()

        porcentaje = (registrados / total) * 100 if total else 0

        return Response({
            "total": total,
            "registrados": registrados,
            "porcentaje": round(porcentaje, 2)
        })

class ReadingGenerationViewSet(TenantSafeMixin,viewsets.ModelViewSet):

    queryset = ReadingGeneration.objects.all()
    serializer_class = ReadingGenerationSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):

        period_str = request.data.get("period")

        if not period_str:
            
            return Response({"error": "Falta el periodo (YYYY-MM)"}, status=400)

        try:

            period_date = datetime.strptime(period_str + "-01", "%Y-%m-%d").date()

        except ValueError:

            return Response({"error": "Formato inválido de periodo"}, status=400)

        # Validar si ya existe una generación para ese periodo
        if ReadingGeneration.objects.filter(period=period_date).exists():

           return Response({"error": f"Ya se generaron lecturas para {period_str}."}, status=400)

        customers = Customer.objects.filter(has_meter=False, state__in=['active', 'inactive'], status=True)

        created = 0
        skipped_existing = 0
        skipped_paid = 0

        for customer in customers:

            # Verificar si ya tiene una lectura para ese periodo
            existing_reading = Reading.objects.filter(customer=customer, period=period_date).first()

            if existing_reading:

                skipped_existing += 1
                continue

            # Verificar si ya tiene una deuda pagada de ese periodo
            if Debt.objects.filter(customer=customer, period=period_date, paid=True).exists():

                skipped_paid += 1
                continue

            tariff = customer.category

            sub_total_amount = tariff.price_water + tariff.price_sewer

            # Crear lectura
            Reading.objects.create(

                customer=customer,
                period=period_date,
                previous_reading=0,
                current_reading=0,
                consumption=0,
                total_water=tariff.price_water,
                total_sewer=tariff.price_sewer,
     
                sub_total_amount=sub_total_amount,
                total_amount=sub_total_amount,
                paid=False,
                date_of_issue=request.data.get("date_of_issue"),
                date_of_due=request.data.get("date_of_due"),
                date_of_cute=request.data.get("date_of_cute")
            )

            created += 1

        # Registrar la generación
        generation = ReadingGeneration.objects.create(
            period=period_date,
            created_by=None,
            total_generated=created,
            notes=request.data.get("notes") or "Generación automática para clientes sin medidor",
            date_of_issue=request.data.get("date_of_issue"),
            date_of_due=request.data.get("date_of_due"),
            date_of_cute=request.data.get("date_of_cute")
        )

        return Response({
            "message": f"Generación completada para {period_str}.",
            "total_creados": created,
            "omitidos_existentes": skipped_existing,
            "omitidos_pagados": skipped_paid
        }, status=201)
    
    @transaction.atomic
    def destroy(self, request, *args, **kwargs):

        instance = self.get_object()
        period = instance.period

        # Filtrar lecturas eliminables
        readings_to_delete = Reading.objects.filter(
            period=period,
        )

        deleted_count = readings_to_delete.count()

        # Eliminar deudas vinculadas a esas lecturas
        debts_to_delete = Debt.objects.filter(
            period=period,
            reading__in=readings_to_delete
        )
        debts_to_delete.delete()

        # Eliminar lecturas
        readings_to_delete.delete()

        # Eliminar la generación
        instance.delete()

        return Response({"message": f"Generación del periodo {period.strftime('%Y-%m')} anulada correctamente.", "lecturas_eliminadas": deleted_count}, status=204)

    # RECIBOS POR CALLE
    @action(detail=False, methods=['get'])
    def download_all_receipts(self, request):
        """
        Descargar un único PDF con todos los recibos de este periodo
        dentro de un ZIP
        """
        company = Company.objects.first()

        month = request.query_params.get("month")
        calle_id = request.query_params.get("calle")
        tenant = request.tenant.schema_name

            # Ruta al logo según el RUC
        logo_path = None

        if company and company.ruc:

            abs_logo_path = os.path.join(settings.MEDIA_ROOT, f"{company.ruc}.jpeg")

            if os.path.exists(abs_logo_path):

               logo_path = Path(abs_logo_path).as_uri()


        if not month:
            return Response({"error": "Debe enviar el parámetro month"}, status=400)

        # 🔹 convertir "2026-04" → date
        try:
            period = datetime.strptime(month, "%Y-%m").date()
        except ValueError:
            return Response({"error": "Formato de month inválido (YYYY-MM)"}, status=400)

        # 🔹 base queryset
        readings = Reading.objects.filter(
            period__year=period.year,
            period__month=period.month
        ).select_related("customer", "customer__calle")

        # conceptos que se deben mostrar en el recibo
        master_concepts = CashConcept.objects.filter(is_master_view=True,state=True).order_by("id")

        # 🔹 filtro por calle (opcional)
        calle = None
        if calle_id:
            calle = Calle.objects.get(pk=calle_id)
            readings = readings.filter(customer__calle_id=calle_id)


        if not readings.exists():
            return Response(
                {"error": "No hay lecturas con calle asignada para este periodo"},
                status=400
            )

        all_readings_context = []

        for reading in readings:
             
            receipt_details = []

            debt = Debt.objects.filter( customer_id=reading.customer_id, period=reading.period).first()

            if not debt:
               
               continue
        
            # convertir detalles de deuda a diccionario
            detail_map = {
                detail.concept_id: detail.amount
                for detail in debt.details.all()
            }

            for concept in master_concepts:

                receipt_details.append({
                    "concept": concept,
                    "amount": detail_map.get(concept.id, 0)
                })
                    # obtener deudas anteriores no pagadas
            previous_debts = Debt.objects.filter(
                customer=reading.customer,
                paid=False,
                period__lt=reading.period
            ).order_by("period")

            # Agrupar por año
            yearly_data = defaultdict(lambda: {"total": 0, "months": []})
            for d in previous_debts:
                year = d.period.year
                month = d.period.month
                yearly_data[year]["total"] += float(d.amount)
                yearly_data[year]["months"].append(month)

            grouped_debts = []
            for year, data in yearly_data.items():
                min_month = min(data["months"])
                max_month = max(data["months"])
                grouped_debts.append({
                    "year": year,
                    "total": f"{data['total']:.2f}",
                    "from_month": format_date(date(year, min_month, 1), "MMMM", locale="es").capitalize(),
                    "to_month": format_date(date(year, max_month, 1), "MMMM", locale="es").capitalize(),
                })

            grouped_debts.sort(key=lambda x: x["year"], reverse=True)

            total_previous_debt = previous_debts.aggregate(total=Sum("amount"))["total"] or 0

            ######################################################
            #           OTROS CONCEPTOS PENDIENTES
            ######################################################

            pending_services = ServiceCharge.objects.filter(
                customer=reading.customer,
                status='pending'
            ).exclude(
                invoice__status='cancelled'
            ).select_related("concept")

            total_other_services = pending_services.aggregate(
                total=Sum("amount")
            )["total"] or 0

            other_services = []

            for service in pending_services:

                other_services.append({
                    "concept": service.concept.name,
                    "description": service.description,
                    "amount": service.amount,
                    "created_at": service.created_at,
                })

            ######################################################
            #           CUOTA PENDIENTE REFINANCIAMIENTO
            ######################################################

            pending_installment = (
                RefinancingInstallment.objects
                .filter(
                    refinancing__customer=reading.customer,
                    paid=False,
                    refinancing__paid=False
                )
                .select_related("refinancing")
                .order_by("due_date", "number")
                .first()
            )

            refinancing_data = None
            total_refinancing = 0

            if pending_installment:

                refinancing_data = {
                    "number": pending_installment.number,
                    "total_amount": pending_installment.total_amount,
                    "due_date": pending_installment.due_date,
                    "refinancing_id": pending_installment.refinancing.id,
                }

                total_refinancing = pending_installment.total_amount

            if debt.paid:

                total_general = (total_previous_debt + total_other_services + total_refinancing)

            else:

                total_general = ( debt.amount + total_previous_debt + total_other_services + total_refinancing)

            total_previous_debt_global = (total_previous_debt +  total_other_services + total_refinancing)

            all_readings_context.append({
                "debt": debt,
                "details": receipt_details,
                "reading": reading,
                "grouped_debts": grouped_debts,
                "total_previous_debt": total_previous_debt,
                "total_other_services": total_other_services,
                "total_previous_debt_global" : total_previous_debt_global,
                "other_services": other_services,
                "refinancing_data": refinancing_data,
                "total_general": total_general,
            })
        
        background_image = None
        template = "agua/recibo.html"

        if tenant == 'chilca':

            background_image = request.build_absolute_uri(f"/media/chilca.png")
            template = "agua/chilca.html"

        # Renderizamos todos los recibos (un reading por página)
        html_content = render_to_string(template, {
            "readings_context": all_readings_context,
            "company": company,
            "company_logo": logo_path,
            "background_image" : background_image
        })

        pdf_bytes = HTML(string=html_content, base_url=request.build_absolute_uri('/')).write_pdf()
        filename = f"recibos_{month}"
        if calle:
           filename = f"{calle.name}_{month}"
        # Devolvemos directamente el PDF
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="{filename}.pdf"'
        )

        return response
 
    # POR ZONA / TICKET
    @action(detail=True, methods=['post'])
    def generate_receipts(self, request, pk=None):

        reading_generation = self.get_object()

        type_ = request.data.get("type", "masivo")
        month = request.data.get("month")
        date_of_issue = request.data.get("date_of_issue")
        date_of_due = request.data.get("date_of_due")
        date_of_cute = request.data.get("date_of_cute")

        updated_count = Reading.objects.filter(
        
            period=reading_generation.period
        
        ).update(
        
            date_of_issue=date_of_issue,
            date_of_due=date_of_due,
            date_of_cute=date_of_cute
        )

        schema_name = request.tenant.schema_name

        if not month:

            raise ValidationError("El campo 'month' es obligatorio")

        try:
            period = datetime.strptime(month, "%Y-%m").date()
        except ValueError:
            raise ValidationError("Formato inválido. Use YYYY-MM")

        description = f"Generación masiva de recibos - Periodo: {period.strftime('%Y-%m')}"
      
        zona_value = None

        if type_ == 'zona':
            zona_pk = request.data.get("zona")

            if not zona_pk:
                raise ValidationError("Debe enviar zona")

            zona = Zona.objects.filter(pk=zona_pk).first()
            if not zona:
                raise ValidationError("Zona no encontrada")

            description = f"Generación de recibos - Zona: {zona.name} - Periodo: {period.strftime('%Y-%m')}"
            zona_value = zona.pk

        elif type_ != "masivo":
            
            raise ValidationError("Tipo inválido")

        # 🔥 IMPORTANTE: transacción para evitar duplicados de ticket
        with transaction.atomic():

            ticket = generate_ticket(period)

            batch = ReceiptBatch.objects.create(
                tenant=request.tenant.schema_name,
                period=period,
                type=type_,
                status="pending",
                description=description,
                ticket=ticket,
                zona_id=zona_value
            )

        async_task(
            "apps.agua.tasks.receipt_tasks.generate_receipts_task",
            batch.id,
            schema_name
        )

        return Response({
            "message": (
                f"Se ha generado un ticket para la generación de recibos. "
                f"El número de ticket es {batch.ticket}. "
                f"Puede verificar el estado en el módulo de Recibos."
            ),
            "ticket": batch.ticket,
            "status": "started"
        })
    
    @action(detail=False, methods=['get'])
    def download_receipts(self, request):

        batch_id = request.query_params.get("batch_id")
        batch = ReceiptBatch.objects.get(id=batch_id)
        schema_name = request.tenant.schema_name

        base_path = os.path.join(
            settings.MEDIA_ROOT,
            "tenants",
            schema_name,
            "recibos",
            batch.ticket,
            str(batch.period)
        )

        zip_path = os.path.join(base_path, f"{batch.ticket}.zip")

        # 🔥 crear zip solo si no existe
        if not os.path.exists(zip_path):

            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:

                for root, dirs, files in os.walk(base_path):
                    for file in files:
                        if file.endswith(".pdf"):
                            full_path = os.path.join(root, file)
                            arcname = os.path.relpath(full_path, base_path)
                            zipf.write(full_path, arcname)

        return FileResponse(open(zip_path, 'rb'), as_attachment=True)

    @action(detail=False, methods=['get'])
    def by_period(self, request):

        period = request.query_params.get("period")

        if not period:
            return Response({"error": "period requerido (YYYY-MM)"}, status=400)

        try:
            year, month = map(int, period.split('-'))
            period_date = date(year, month, 1)
        except:
            return Response({"error": "Formato inválido"}, status=400)

        generation = ReadingGeneration.objects.filter(period=period_date).first()

        if not generation:
            return Response({
                "exists": False
            })

        return Response({
            "exists": True,
            "id": generation.id,
            "period": generation.period,
            "status": "open",  # luego puedes hacerlo dinámico
            "total_generated": generation.total_generated
        })

class DebtViewSet(TenantSafeMixin,viewsets.ModelViewSet):

    queryset = Debt.objects.all().order_by('period')
    serializer_class = DebtSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = DebtFilter

    @transaction.atomic
    def create(self, request, *args, **kwargs):

        tenant = request.tenant.schema_name
        data = request.data

        customer_id = data.get("customer")
        period_str = data.get("period")

        # Validaciones básicas
        if not customer_id or not period_str:
            raise ValidationError(
                "Debe enviar 'customer' y 'period'."
            )

        # Normalizar periodo al primer día del mes
        try:

            period = date.fromisoformat(period_str)

        except ValueError:

            raise ValidationError(
                "El formato del periodo es inválido."
            )

        normalized_period = date(
            period.year,
            period.month,
            1
        )

        # Obtener cliente
        try:

            customer = Customer.objects.select_related(
                "category"
            ).get(id=customer_id)

        except Customer.DoesNotExist:

            raise ValidationError(
                "El cliente no existe."
            )

        # Evitar duplicados
        if Debt.objects.filter(
            customer=customer,
            period=normalized_period
        ).exists():

            raise ValidationError(
                "Ya existe una deuda para este cliente y periodo."
            )

        # =========================
        # MONTOS BASE
        # =========================

        total_water = customer.category.price_water or Decimal("0.00")
        total_sewer = customer.category.price_sewer or Decimal("0.00")

        billing_type = customer.billing_type or "both"

        # Ajustar según tipo de facturación
        if billing_type == "water":

            total_sewer = Decimal("0.00")

        elif billing_type == "sewer":

            total_water = Decimal("0.00")

        # =========================
        # CONCEPTOS ADICIONALES
        # =========================

        price_clean = get_concept_total("price_clean")
        price_fixed_charge = get_concept_total("price_fixed_charge")
        price_maintenance = get_concept_total( "price_maintenance")

        # =========================
        # LÓGICA PERSONALIZADA TENANT
        # =========================

        if tenant == "pangoa":

            if customer.state == "inactive":

                total_water = Decimal("0.00")
                total_sewer = Decimal("0.00")

            else:

                price_fixed_charge = 0
                price_maintenance = 0

        # =========================
        # TOTALES
        # =========================

        total_amount_reading = (
            total_water +
            total_sewer
        )

        total_amount_debt = (
            total_amount_reading +
            price_clean +
            price_fixed_charge +
            price_maintenance
        )

        # =========================
        # CREAR LECTURA
        # =========================

        reading = None

        # Crear lectura SOLO si tiene no medidor
        if not customer.has_meter:

            reading = Reading(
                customer=customer,
                period=normalized_period,
                current_reading=Decimal("0.000"),
                has_meter=customer.has_meter,
                total_water=total_water,
                total_sewer=total_sewer,
                sub_total_amount=total_amount_reading,
                total_amount=total_amount_reading,
            )

            reading.save(skip_process=True)

        # =========================
        # CREAR DEUDA
        # =========================

        debt = Debt.objects.create(
            customer=customer,
            period=normalized_period,
            amount=total_amount_debt,
            description=(
                f"Deuda del periodo "
                f"{period.strftime('%Y-%m')}"
            ),
            reading=reading,
        )

        # =========================
        # DETALLE DE CONCEPTOS
        # =========================

        concept_map = {

            "price_water": total_water,
            "price_sewer": total_sewer,

            "price_clean": price_clean,
            "price_fixed_charge": price_fixed_charge,
            "price_maintenance": price_maintenance,
        }

        concepts = {
            concept.system_key: concept
            for concept in CashConcept.objects.filter(
                system_key__in=concept_map.keys()
            )
        }

        for system_key, amount in concept_map.items():

            if amount > 0:

                DebtDetail.objects.create(
                    debt=debt,
                    concept=concepts[system_key],
                    amount=amount
                )

        # =========================
        # RESPUESTA
        # =========================

        serializer = self.get_serializer(debt)

        return Response(
            serializer.data,
            status=status.HTTP_201_CREATED
        )
    
    @transaction.atomic
    def update(self, request, *args, **kwargs):

        instance = self.get_object()
        data = request.data

        details_data = data.get("details", [])

        sent_ids = [d.get("id") for d in details_data if d.get("id")]

        # eliminar los detalles que ya no vienen
        for detail in instance.details.all():

            if detail.id not in sent_ids:
                detail.delete()

        for d in details_data:

            detail_id = d.get("id")
            concept_id = d.get("concept_id")
            amount = d.get("amount")
            
            detail = DebtDetail.objects.get(id=detail_id, debt=instance)
            detail.concept_id = concept_id or detail.concept_id
            detail.amount = amount
            detail.save()

        # recalcular total
        total = sum(detail.amount for detail in instance.details.all())
      
        instance.amount = total
        instance.save()

        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No se proporciono un archivo.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(
                file,
                engine='openpyxl',
                # header=2,
                dtype={'Codigo': str}
            )

            df['Codigo'] = df['Codigo'].astype(str).str.strip().str[-5:].str.zfill(5)
        except Exception as e:
            return Response({'error': f'Error al leer el archivo: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        errores = []
        procesados = 0

        # 🔹 Precargar conceptos
        conceptos = {
            "001": CashConcept.objects.get(code="001"),
            "002": CashConcept.objects.get(code="002"),
            # "003": CashConcept.objects.get(code="003"),
            # "004": CashConcept.objects.get(code="004"),
            # "005": CashConcept.objects.get(code="005"),
        }

        # 🔹 Precargar clientes del Excel
        codigos_excel = df['Codigo'].unique()
        clientes = {
            c.codigo: c
            for c in Customer.objects.filter(codigo__in=codigos_excel)
        }

        debts_to_create = []
        details_to_create = []

        try:

            cargo_fijo = CashConcept.objects.get(code="003")

        except CashConcept.DoesNotExist:

            cargo_fijo = None

        total_fixed_charge = cargo_fijo.total if cargo_fijo else Decimal("0.00")
        # df = df.head(2)

        for row in df.itertuples(index=False):

            codigo = str(row.Codigo)

  

            year = row.anio
            meses_texto = to_none_if_empty(row.Meses)
            total = to_decimal_or_none(row.Agua)

            if year != 2026:
                if not meses_texto:
                    errores.append({"codigo": codigo, "anio": year, "total": total, "error": "Campo 'Meses' vacio"})
                    continue

                customer = clientes.get(codigo)
                if not customer:
                    errores.append({"codigo": codigo, "anio": year, "meses": meses_texto, "total": total, "error": "Cliente no encontrado"})
                    continue

                try:
                    periodos = generar_periodos(int(year), meses_texto)
                except Exception as e:
                    errores.append({"codigo": codigo, "anio": year, "meses": meses_texto, "total": total, "error": f"Error al generar periodos: {str(e)}"})
                    continue

                # Calcular montos con precisión decimal
                total_water = (Decimal(total) / Decimal(len(periodos))) if (total and len(periodos) > 0) else Decimal("0.00")
                total_sewer = Decimal(customer.category.price_sewer or 0)
                total_fixed_charge = Decimal(customer.category.price_fixed_charge or 0)

                amount = total_water + total_sewer + total_fixed_charge
                if amount <= Decimal("0.00"):
                    print(f"Deuda ignorada para {codigo} monto 0")
                    continue
                for periodo in periodos:
                    # 🔹 Obtener o crear debt en memoria, no en DB aún

                    reading = Reading(
                        customer=customer,
                        period=periodo,
                        current_reading=Decimal("0.000"),
                        has_meter=customer.has_meter,
                        total_water=total_water,
                        total_sewer=total_sewer,
                        total_fixed_charge=total_fixed_charge,
                        total_amount=amount,
                    )
                    reading.save(skip_process=True)

                    debt, created = Debt.objects.get_or_create(
                        customer=customer,
                        reading=reading,
                        period=periodo,
                        defaults={
                            "description": "Deuda importada desde Excel",
                            "amount": amount,
                            "paid": False
                        }
                    )

                    if not created:
                        debt.amount = amount
                        debt.save()
                        debt.details.all().delete()

                    # 🔹 Preparar detalles para bulk_create
                    if total_water > 0:
                        details_to_create.append(DebtDetail(debt=debt, concept=conceptos["001"], amount=total_water))
                    if total_sewer > 0:
                        details_to_create.append(DebtDetail(debt=debt, concept=conceptos["002"], amount=total_sewer))
                    # if total_fixed_charge > 0:
                    #     details_to_create.append(DebtDetail(debt=debt, concept=conceptos["003"], amount=total_fixed_charge))

                    procesados += 1

        # 🔹 Insertar todos los detalles de una vez
        if details_to_create:
            DebtDetail.objects.bulk_create(details_to_create)

        return Response({
            "procesados": procesados,
            "errores": errores
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"])
    def create_reading(self, request, pk=None):
        debt = self.get_object()

        if debt.reading:

            return Response(
                {"detail": "Esta deuda ya tiene una lectura vinculada."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Crear lectura SIN procesar
        reading = Reading(
            customer=debt.customer,
            paid=debt.paid,
            period=debt.period,
            current_reading=Decimal("0.000"),
            has_meter=debt.customer.has_meter,
        )

        reading.save(skip_process=True)  # 👈 aquí usamos el flag

        # Vincular lectura con deuda existente
        debt.reading = reading
        debt.save(update_fields=["reading"])

        return Response(
            {"detail": f"Lectura creada y vinculada correctamente a la deuda {debt.id}."},
            status=status.HTTP_201_CREATED
        )

    @action(detail=False, methods=['post'])
    def refinanciar(self, request):

        tenant = request.tenant.schema_name

        customer_id = request.data.get("customer_id")

        service_id = request.data.get("service_id")

        years = request.data.get("years", [])

        cuotas = int(request.data.get("cuotas", 1))

        type = request.data.get("type", "debt")

        if cuotas <= 0:

            raise ValidationError("Cuotas invalidas")

        total = Decimal('0')

        debts = Debt.objects.none()
        service_charges = ServiceCharge.objects.none()

        ####################################################
        # DEUDAS
        ####################################################

        if type in ['debt', 'all']:

            debts = Debt.objects.filter(
                customer_id=customer_id,
                paid=False,
                is_refinanced=False,
                period__year__in=years
            )

            if not debts.exists():
                raise ValidationError(
                    "No hay deudas validas"
                )

            total = (
                debts.aggregate(total=Sum('amount'))['total']
                or Decimal('0')
            )

        ####################################################
        # SERVICE CHARGES
        ####################################################

        if type in ['service', 'all']:

            service_charges = ServiceCharge.objects.filter(
                pk = service_id,
                # customer_id=customer_id,
                status='pending',
                is_refinanced=False
            )

            total += (
                service_charges.aggregate(total=Sum('amount'))['total']
                or Decimal('0')
            )

        ####################################################
        # INTERES
        ####################################################

        INTEREST_RATE = Decimal('0')

        if tenant == 'chilca':

            INTEREST_RATE = Decimal('0.002')

        total_interest_rate = (
            INTEREST_RATE * cuotas
        )

        interest_amount = (
            total * total_interest_rate
        ).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

        total_with_interest = (
            total + interest_amount
        ).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

        ####################################################
        # CUOTAS BASE
        ####################################################

        capital_per_installment = (
            total / cuotas
        ).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

        interest_per_installment = (
            interest_amount / cuotas
        ).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

        total_installment = (
            capital_per_installment +
            interest_per_installment
        ).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

        with transaction.atomic():

            ####################################################
            # REFINANCIACION
            ####################################################

            ref = DebtRefinancing.objects.create(

                customer_id=customer_id,

                total_amount=total,

                interest_rate=INTEREST_RATE,

                interest_amount=interest_amount,

                total_amount_with_interest=total_with_interest,

                installments=cuotas,

                type = type

            )

            ####################################################
            # VINCULAR DEUDAS
            ####################################################

            for d in debts:

                DebtRefinancingDetail.objects.create(
                    refinancing=ref,
                    debt=d
                )

                d.is_refinanced = True

                d.save()

            ####################################################
            # VINCULAR SERVICE CHARGES
            ####################################################

            for s in service_charges:

                ServiceRefinancingDetail.objects.create(
                    refinancing=ref,
                    service_charge=s,
                )

                s.is_refinanced = True
                s.status = 'refinanced'
                s.save()

            ####################################################
            # GENERAR CRONOGRAMA
            ####################################################

            generated_total = Decimal('0')

            for i in range(1, cuotas + 1):

                due_date = (
                    now().date() +
                    relativedelta(months=i)
                )

                ############################################
                # AJUSTAR ULTIMA CUOTA
                ############################################

                if i == cuotas:

                    cuota_total = (
                        total_with_interest -
                        generated_total
                    ).quantize(
                        Decimal('0.01'),
                        rounding=ROUND_HALF_UP
                    )

                else:

                    cuota_total = total_installment

                    generated_total += cuota_total

                ############################################
                # CAPITAL AJUSTADO
                ############################################

                capital = (
                    cuota_total -
                    interest_per_installment
                ).quantize(
                    Decimal('0.01'),
                    rounding=ROUND_HALF_UP
                )

                RefinancingInstallment.objects.create(

                    refinancing=ref,

                    number=i,

                    capital_amount=capital,

                    interest_amount=interest_per_installment,

                    total_amount=cuota_total,

                    due_date=due_date

                )

        return Response({

            "message": "Refinanciacion creada correctamente",

            "subtotal": total,

            "interest_amount": interest_amount,

            "total": total_with_interest,

            "cuotas": cuotas

        })

    @action(detail=False, methods=['post'])
    def preview_refinanciamiento(self, request):

        tenant = request.tenant.schema_name
    
        customer_id = request.data.get("customer_id")

        service_id = request.data.get("service_id")

        years = request.data.get("years", [])

        cuotas = int(request.data.get("cuotas", 1))

        type = request.data.get("type", "debt")

        if cuotas <= 0:

            raise ValidationError("Cuotas invalidas")

        total = Decimal('0')

        debts = Debt.objects.none()
        service_charges = ServiceCharge.objects.none()

        ####################################################
        # DEUDAS
        ####################################################

        if type in ['debt', 'all']:

            debts = Debt.objects.filter(
                customer_id=customer_id,
                paid=False,
                is_refinanced=False,
                period__year__in=years
            )

            total = (
                debts.aggregate(total=Sum('amount'))['total']
                or Decimal('0')
            )

        ####################################################
        # SERVICE CHARGES
        ####################################################

        if type in ['service', 'all']:

            service_charges = ServiceCharge.objects.filter(
                pk = service_id,
                # customer_id=customer_id,
                status='pending',
                is_refinanced=False
            )

            total += (
                service_charges.aggregate(total=Sum('amount'))['total']
                or Decimal('0')
            )

        ####################################################
        # INTERES
        ####################################################

        INTEREST_RATE = Decimal('0')

        if tenant == 'chilca':

            INTEREST_RATE = Decimal('0.002')


        total_interest_rate = (
            INTEREST_RATE * cuotas
        )

        interest_amount = (
            total * total_interest_rate
        ).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

        total_with_interest = (
            total + interest_amount
        ).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

        ####################################################
        # CUOTAS BASE
        ####################################################

        capital_per_installment = (
            total / cuotas
        ).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

        interest_per_installment = (
            interest_amount / cuotas
        ).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

        total_installment = (
            capital_per_installment +
            interest_per_installment
        ).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

        cuotas_preview = []

        generated_total = Decimal('0')

        ####################################################
        # CRONOGRAMA
        ####################################################

        for i in range(1, cuotas + 1):

            due_date = (
                now().date() +
                relativedelta(months=i)
            )

            ############################################
            # AJUSTAR ULTIMA CUOTA
            ############################################

            if i == cuotas:

                cuota_total = (
                    total_with_interest -
                    generated_total
                ).quantize(
                    Decimal('0.01'),
                    rounding=ROUND_HALF_UP
                )

            else:

                cuota_total = total_installment

                generated_total += cuota_total

            ############################################
            # CAPITAL AJUSTADO
            ############################################

            capital = (
                cuota_total -
                interest_per_installment
            ).quantize(
                Decimal('0.01'),
                rounding=ROUND_HALF_UP
            )

            cuotas_preview.append({

                "numero": i,

                "fecha_vencimiento": due_date,

                "capital": capital,

                "interes": interest_per_installment,

                "total": cuota_total

            })

        return Response({

            "subtotal": total,

            "interest_rate": INTEREST_RATE,

            "interest_amount": interest_amount,

            "total": total_with_interest,

            "cuotas": cuotas_preview

        })

    @action(detail=True, methods=['post'])
    def toggle_paid(self, request, pk=None):

        debt = self.get_object()

        # Cambiar solo paid
        debt.paid = not debt.paid
        debt.save(update_fields=['paid'])

        # Sin recalcular nada
        if debt.reading:

            debt.reading.paid = debt.paid

            debt.reading.save(
                skip_process=True,
                update_fields=['paid']
            )

        return Response({
            "success": True,
            "paid": debt.paid
        })

class RefinancingInstallmentViewSet(TenantSafeMixin,viewsets.ModelViewSet):

    queryset = RefinancingInstallment.objects.all()
    serializer_class = RefinancingInstallmentSerializer
  
    def get_queryset(self):

        queryset = super().get_queryset()

        customer_id = self.request.query_params.get('customer')
        paid = self.request.query_params.get('paid')

        if customer_id:
            queryset = queryset.filter(
                refinancing__customer_id=customer_id  # 🔥 clave
            )

        if paid is not None:
            queryset = queryset.filter(paid=(paid == 'true'))

        return queryset.order_by('number')

class InvoiceViewSet(TenantSafeMixin, viewsets.ModelViewSet):

    queryset = Invoice.objects.all().order_by('-id')
    serializer_class = InvoiceSerializer
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend,filters.SearchFilter]
    filterset_fields = ['status']  
    search_fields = ['customer__codigo', 'customer__full_name', 'customer__number','code']

    permission_classes = [TenantPaymentCreatePermission]

    @action(detail=True, methods=['get'], url_path='ticket')
    def ticket_pdf(self, request, pk=None, **kwargs):

        invoice = get_object_or_404(Invoice, id=pk)

        # Usamos la relación inversa para evitar consultas innecesarias
        payments_debts = invoice.invoice_debts.select_related('debt').order_by('debt__period')
        payments_concepts = invoice.invoice_concepts.select_related('concept').order_by('concept__code')
        payments_installments = invoice.invoice_installments.select_related('installment__refinancing').order_by('installment__number')

        company = Company.objects.first()


        # Ruta al logo según el RUC
        logo_path = None
        if company and company.ruc:
            logo_path = request.build_absolute_uri(f"/media/{company.ruc}.jpeg")

        print(invoice.reference)

        context = {
            "invoice": invoice,
            "customer": invoice.customer,
            "concepts": payments_concepts,
            "payments": payments_debts,
            "installments": payments_installments,  # 👈 NUEVO
            "total_paid": sum((p.total for p in payments_debts), 0),
            "total_paid_concept": sum((p.total for p in payments_concepts), 0),
            # "total_paid_installments": sum((p.total for p in payments_installments), 0),  # 👈 opcional
            "company_name": company.name if company else "",
            "company_ruc": company.ruc if company else "",
            "company_logo": logo_path,
        }
        
        template = get_template('agua/invoice.html')
        html_string = template.render(context)

        pdf_buffer = io.BytesIO()
   
        HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(
            pdf_buffer
        )

        file_name = f"ticket_{invoice.id}.pdf"
        pdf_buffer.seek(0)
        response = HttpResponse(pdf_buffer.read(), content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{file_name}"'
        return response

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        invoice = self.get_object()
        invoice.cancel()
        return Response({"message": "Factura anulada"}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def preview_sheets(self, request):

        file = request.FILES.get("file")

        excel_file = pd.ExcelFile(file)

        return Response({
            "sheets": excel_file.sheet_names
        })

class CashConceptViewSet(TenantSafeMixin, viewsets.ModelViewSet):

    queryset = CashConcept.objects.all().order_by('id')
    serializer_class = CashConceptSerializer

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter
    ]

    search_fields = ['name']

    filterset_fields = ['is_master', 'is_master_view', 'state']

class CategoryViewSet(TenantSafeMixin,viewsets.ModelViewSet):
    
    permission_classes = [IsAuthenticated]
    serializer_class = CategorySerializer
    queryset = Category.objects.all() 

    def get_queryset(self):

        return Category.objects.filter(state=True).order_by('id')

    @action(detail=False, methods=['post'])
    def import_excel(self, request):

        file = request.FILES.get('file')

        if not file:

            return Response({'error': 'No se proporciono un archivo.'}, status=status.HTTP_400_BAD_REQUEST)

        try:

            extension = os.path.splitext(file.name)[1].lower()

            if extension == ".xls":

                df = pd.read_excel(file, engine="xlrd", dtype={'codigo': str})

            elif extension == ".xlsx":

                df = pd.read_excel(file, engine="openpyxl", dtype={'codigo': str})

            else:

                return Response({'error': 'Formato no soportado. Solo .xls o .xlsx'}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:

            return Response({'error': f'Error al leer el archivo: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        for index, row in df.sort_values(by='codigo').iterrows():

            codigo = str(row.get('codigo')).zfill(2)  # Siempre 2 dígitos
            descrip = row.get('descrip')
            agua = row.get('agua')

            # Crear o actualizar registro
            Category.objects.update_or_create(
                codigo=codigo,
                defaults={
                    'name': descrip,
                    'price_water': agua,
                    'price_sewer': 0,
                    'has_meter': False  
                }
            )

            print(f"Importado: {codigo} - {descrip} - {agua}")


        return Response({"message":"ubicacion cargada"}, status=status.HTTP_200_OK)

class ViaViewSet(TenantSafeMixin,viewsets.ModelViewSet):

    queryset = Via.objects.all().order_by('id')
    serializer_class = ViaSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name']
    ordering_fields = ['name']

    @action(detail=False, methods=['post'])
    def import_excel(self, request):

        file = request.FILES.get('file')

        if not file:

            return Response({'error': 'No se proporciono un archivo.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
        
            df = pd.read_excel(
                    file,
                    engine='openpyxl',
                    dtype={
                        'tipo_dir': str,
                        'codigo': str
                    })

        except Exception as e:

            return Response({'error': f'Error al leer el archivo: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # for index, row in df.sort_values(by='tipo_dir').iterrows():

        #     name = row.get('abrv')
        #     codigo = str(row.get('tipo_dir')).zfill(2)  # Siempre 2 dígitos

        #     if Via.objects.filter(codigo=codigo).exists():
        #         continue

        #     via = Via(name=name, codigo=codigo)
        #     via.save()

        df = df.sort_values(by=['codigo'], ascending=True)
        for index, row in df.iterrows():

            codigo = str(row.get('codigo') or '').strip()
            name = str(row.get('nombre') or '').strip()
            codigo_via = int(row.get('tipo_dir'))

            print(codigo, name, codigo_via)

            if not name or not codigo_via:
                
               print(f'Fila {index + 2}: calle invalida (nombre o id_via vacio)')
               continue

            try:

                via = Via.objects.get(pk=codigo_via)

            except Via.DoesNotExist:

                print(f'Fila {index + 2}: via con codigo {codigo_via} no existe (para la calle "{name}")')

                continue

            # if Calle.objects.filter(name=name, via=via).exists():

            #     print(f'Fila {index + 2}: ya existe la calle "{name}" en la via {via.name}')
            #     continue

            calle = Calle(name=name, via=via, codigo=codigo)
            calle.save()

        return Response({"message":"ubicacion cargada"}, status=status.HTTP_200_OK)

class CalleViewSet(TenantSafeMixin,viewsets.ModelViewSet):

    queryset = Calle.objects.select_related('via').all().order_by('id')
    serializer_class = CalleSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['via']  # permite filtrar por tipo_via id
    search_fields = ['codigo','name']

class CompanyViewSet(TenantSafeMixin,viewsets.ModelViewSet):

    queryset = Company.objects.all()
    serializer_class = CompanySerializer

class CashOutflowViewSet(TenantSafeMixin,viewsets.ModelViewSet):

    queryset = CashOutflow.objects.all().order_by('-id')
    serializer_class = CashOutflowSerializer
    pagination_class = CustomPagination

class ConfigViewSet(TenantSafeMixin,viewsets.ModelViewSet):

    authentication_classes = []
    permission_classes = []

    queryset = Config.objects.all().order_by('-id')
    serializer_class = ConfigSerializer
    pagination_class = CustomPagination

class MorosidadViewSet(TenantSafeMixin, viewsets.ModelViewSet):

    serializer_class = MorosidadSerializer
    pagination_class = CustomPagination

    @action(detail=False, methods=["get"], url_path="moroso")
    def overdue(self, request):

        zona_id = request.query_params.get("zona")
        min_months = int(request.query_params.get("min_months", 1))

        queryset = get_morosos_queryset(zona_id, min_months).order_by('codigo')

        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(page, many=True)
        return self.get_paginated_response(serializer.data)

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):

        zona_id = request.query_params.get("zona")
        min_months = int(request.query_params.get("min_months", 1))

        queryset = get_morosos_queryset(zona_id, min_months).order_by('codigo')

        data = []

        for obj in queryset:
       
            data.append({
                "Código": obj.codigo,
                "Cliente": obj.full_name,
                "Dirección": obj.address,
                "Meses Deuda": obj.unpaid_months or 0,
                "Total Deuda": obj.total_debt or 0,
                "Corte Pendiente": "Sí" if obj.has_pending_cut else "No",
                "Corte Ejecutado": "Sí" if obj.has_executed_cut else "No",
            })

        df = pd.DataFrame(data)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="morosidad.xlsx"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Morosidad')

            ws = writer.sheets['Morosidad']

            # 🔥 AUTOAJUSTAR COLUMNAS
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter

                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                ws.column_dimensions[col_letter].width = max_length + 2

            # 🔥 HEADER EN NEGRITA
            from openpyxl.styles import Font
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # 🔥 FORMATO MONEDA
            for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
                for cell in row:
                    cell.number_format = '#,##0.00'

            # 🔥 FORMATO FECHA
            for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
                for cell in row:
                    cell.number_format = 'YYYY-MM-DD'

        return response

    @action(detail=False, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request):

        zona_id = request.query_params.get("zona", None)
        min_months = int(request.query_params.get("min_months", 1))

        MAX_ROWS = 1000
        
        zona = None

        if zona_id:

            try:

                zona = Zona.objects.get(pk=zona_id)

            except Zona.DoesNotExist:

                return Response({
                    "error": "La zona no existe."
                }, status=status.HTTP_404_NOT_FOUND)

        queryset = get_morosos_queryset(zona_id, min_months).order_by('codigo')

        # ✅ TOTAL REAL
        total = 1000

        # 🚨 VALIDACIÓN REAL
        if total > MAX_ROWS:
            return Response({
                "error": f"El reporte PDF excede el límite ({MAX_ROWS} registros).",
                "suggestion": "Use Excel para grandes volúmenes."
            }, status=status.HTTP_400_BAD_REQUEST)

        # ✅ SIN CORTE OCULTO
        data = [
            {
                "codigo": obj.codigo,
                "nombre": obj.full_name,
                "address" : obj.address,
                "meses": obj.unpaid_months or 0,
                "deuda": obj.total_debt or 0,
                "pendiente": "Sí" if obj.has_pending_cut else "No",
                "ejecutado": "Sí" if obj.has_executed_cut else "No",
            }
            for obj in queryset[:1000]
        ]

        html_string = render_to_string(
            "morosidad/morosidad_report.html",
            {
                "data": data,
                "total": total,
                "zona": zona.name if zona else "Todas",
                "min_months": min_months,
                "fecha": now().strftime("%d/%m/%Y %H:%M")
            }
        )

        response = HttpResponse(content_type='application/pdf')

        # 👇 puedes elegir
        response['Content-Disposition'] = 'attachment; filename="morosidad.pdf"'
        # response['Content-Disposition'] = 'inline; filename="morosidad.pdf"'

        HTML(string=html_string).write_pdf(response)

        return response

class CutBatchViewSet(TenantSafeMixin, viewsets.ModelViewSet):

    queryset = CutBatch.objects.all().order_by('-id')
    serializer_class = CutBatchSerializer
    pagination_class = CustomPagination

    @action(detail=True, methods=["get"])
    def cuts(self, request, pk=None):

        batch = self.get_object()

        cuts = batch.cuts.select_related('customer').order_by(
            'customer__sector',
            'customer__calle',
            'customer__nro'
        )

        serializer = ServiceCutSerializer(cuts, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def export_pdf(self, request, pk=None):

        batch = self.get_object()

        cuts = (
            ServiceCut.objects
            .filter(batch=batch)
            .select_related(
                "customer",
                "customer__calle",
                "customer__zona",
                "customer__manzana"
            )
            .prefetch_related(
                Prefetch(
                    "customer__meterassignment_set",
                    queryset=MeterAssignment.objects.filter(
                        is_active=True
                    ).select_related("meter"),
                    to_attr="active_assignments"
                )
            )
            .annotate(

                # ==========================================
                # TOTALES
                # ==========================================

                total_debt=Sum(
                    "customer__debts__amount",
                    filter=Q(customer__debts__paid=False)
                ),

                unpaid_months=Count(
                    "customer__debts__id",
                    filter=Q(customer__debts__paid=False),
                    distinct=True
                ),

                # ==========================================
                # ORDENAMIENTO
                # ==========================================

                mz_number=Cast(
                    "customer__manzana__codigo",
                    IntegerField()
                ),

                predio_number=Cast(
                    "customer__predio",
                    IntegerField()
                ),
            )
            .prefetch_related("debts")
            .order_by(
                "customer__sector",
                "mz_number",
                "predio_number",
            )
        )

        html_string = render_to_string("cut_batch/cut_batch.html", {
            "cuts": cuts,
            "date": batch.scheduled_date,
            "zona": batch.sector or "-",
            "batch": batch
        })

        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf = html.write_pdf()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="padron_corte_{batch.id}.pdf"'

        return response

class ServiceCutViewSet(TenantSafeMixin,viewsets.ModelViewSet):

    queryset = ServiceCut.objects.all()
    serializer_class = ServiceCutSerializer
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend,filters.SearchFilter]

    filterset_fields = ['status']  

    def _build_cut_queryset(self, status):

        return (
            self.filter_queryset(self.get_queryset())
            .filter(status=status)
            .select_related('customer')
            .prefetch_related('debts')
            .order_by('customer__codigo')
        )
 
    #  GENERAR PADRÓN
    @action(detail=False, methods=["post"])
    def create_batch(self, request):

        zone = request.data.get("zona")
        name = request.data.get("name")
        sector = request.data.get("sector")
        scheduled_date = request.data.get("scheduled_date")
        min_months = int(request.data.get("min_months", 1))
 
        with transaction.atomic():

            queryset = get_morosos_queryset(zone, min_months,'active')

            customers = queryset.prefetch_related(
                Prefetch(
                    'debts',
                    queryset=Debt.objects.filter(paid=False),
                    to_attr='pending_debts'
                )
            )

            if not customers.exists():

                return Response({
                    "error": "No se encontrado clientes"
                }, status=404)


            customer_ids = list(customers.values_list('id', flat=True))

            existing_customers = set(
                ServiceCut.objects.filter(
                    customer_id__in=customer_ids,
                    status="pending"
                ).values_list("customer_id", flat=True)
            )

            # 🔥 filtrar clientes válidos
            valid_customers = [
                customer for customer in customers
                if customer.id not in existing_customers
            ]

            # ❌ evitar batch vacío
            if not valid_customers:
                return Response({ "error": "Todos los clientes ya estan en proceso"}, status=400)

            # ✅ recién aquí creas el batch
            batch = CutBatch.objects.create(
                name=name,
                zone=zone,
                sector=sector,
                scheduled_date=scheduled_date
            )

            created = []

            for customer in customers:

                if customer.id in existing_customers:
                    
                    continue

                debts = customer.pending_debts

                cut = ServiceCut.objects.create(
                    customer=customer,
                    scheduled_date=scheduled_date,
                    batch=batch,
                    created_by=request.user.id
                )

                cut.debts.set(debts)
                created.append(cut.id)

        return Response({
            "batch_id": batch.id,
            "created": len(created)
        })
  
    @action(detail=True, methods=["post"])
    def execute(self, request, pk=None):

        cut = self.get_object()

        # validar estado
        if cut.status != "pending":
            return Response({
                "error": "El corte no está pendiente"
            }, status=400)

        try:
            cut.execute_cut(
                user_id=request.user.id,
                result=request.data.get("result", "executed"),
                observation=request.data.get("observation")
            )

            # 🔥 actualizar estado del batch
            batch = cut.batch
            if batch:
                pending_exists = batch.cuts.filter(status="pending").exists()

                if pending_exists:
                    batch.status = "in_progress"
                else:
                    batch.status = "completed"

                batch.save()

            # 🔥 devolver instancia actualizada
            serializer = self.get_serializer(cut)

            return Response(serializer.data)

        except Exception as e:
            return Response({"error": str(e)}, status=500)

    @action(detail=False, methods=["get"], url_path="report-excel")
    def report_excel(self, request):

        status_param = request.query_params.get("status")
        print(status_param)

        queryset = self._build_cut_queryset(status_param)

        data = []

        for obj in queryset:

            data.append({
                "Código": obj.customer.codigo,
                "Cliente": obj.customer.full_name,
                "Dirección": obj.customer.address,
                "Sector": obj.customer.sector,
                "Estado Corte": obj.status,
                "Motivo": obj.reason,
                "Fecha Programada": obj.scheduled_date,
                "Fecha Ejecución": obj.execution_date,
            })

        df = pd.DataFrame(data)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        filename = f"cortes_{status_param or 'todos'}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Cortes')

            ws = writer.sheets['Cortes']

            # auto width
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2

        return response

    @action(detail=False, methods=["get"], url_path="report-pdf")
    def report_pdf(self, request):

        MAX_ROWS = 1000
        status_param = request.query_params.get("status")
        status_display = dict(ServiceCut.STATUS_CHOICES).get(status_param, status_param)

        queryset = self._build_cut_queryset(status_param)

        total = queryset.count()

        if total > MAX_ROWS:
            return Response({
                "error": f"El PDF excede el límite ({MAX_ROWS})",
                "suggestion": "Use Excel para grandes volúmenes"
            }, status=status.HTTP_400_BAD_REQUEST)

        data = [
            {
                "codigo": obj.customer.codigo,
                "nombre": obj.customer.full_name,
                "direccion": obj.customer.address,
                "sector": obj.customer.sector,
                "estado_cliente": obj.customer.state,
                "estado_corte": obj.get_status_display(),
                "motivo": obj.reason,
                "fecha_prog": obj.scheduled_date,
                "fecha_ejec": obj.execution_date,
                "deudas": obj.debts.count(),
            }
            for obj in queryset
        ]

        html_string = render_to_string("service_cut/report.html", {
            "data": data,
            "total": total,
            "status": status_display  
        })

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="cortes_{status_param}.pdf"'

        HTML(string=html_string).write_pdf(response)

        return response

class DashboardSummaryAPIView(TenantSafeMixin, APIView):

    def get(self, request):

        today = date.today()
        month_start = today.replace(day=1)

        # 1️⃣ Clientes totales
        total_customers = Customer.objects.count()

        # 2️⃣ Clientes con deuda
        customers_with_debt = (
            Debt.objects
            .filter(paid=False)
            .values('customer')
            .distinct()
            .count()
        )

        # 3️⃣ Recaudación del mes
        monthly_revenue = (
            InvoicePayment.objects
            .filter(created_at__date__gte=month_start)
            .aggregate(total=Sum('total'))['total'] or 0
        )

        # 4️⃣ Recaudación del día
        daily_revenue = (
            InvoicePayment.objects
            .filter(created_at__date=today)
            .aggregate(total=Sum('total'))['total'] or 0
        )

        # 5️⃣ Top 5 clientes con más deuda
        top_debtors = (
            Debt.objects
            .filter(paid=False)
            .values(
                'customer__id',
                'customer__full_name'
            )
            .annotate(total_debt=Sum('amount'))
            .order_by('-total_debt')[:5]
        )

        weekly_income = []

        for i in range(4, 0, -1):
            start_week = today - timedelta(days=i * 7)
            end_week = start_week + timedelta(days=6)

            total = (
                InvoicePayment.objects
                .filter(
                    created_at__date__gte=start_week,
                    created_at__date__lte=end_week
                )
                .aggregate(total=Sum('total'))['total'] or 0
            )

            weekly_income.append({
                "label": f"Semana {5 - i}",
                "total": float(total)
            })

        # 6️⃣ Deuda por zona
        debt_by_zone = (
            Debt.objects
            .filter(paid=False, customer__zona__isnull=False)
            .values(
                'customer__zona__id',
                'customer__zona__name'
            )
            .annotate(total_debt=Sum('amount'))
            .order_by('-total_debt')
        )

        # Total deuda pendiente
        total_debt = (
            Debt.objects
            .filter(paid=False)
            .aggregate(total=Sum('amount'))['total'] or 0
        )

        # Total recaudado (histórico)
        total_collected = (
            Invoice.objects
            .aggregate(total=Sum('total'))['total'] or 0
        )


        return Response({
            "cards": [
                {
                    "title": "Clientes Totales",
                    "icono" : "fa-solid fa-users",
                    "value": total_customers,
                },
                {
                    "title": "Clientes con Deuda",
                    "icono" : "fa-solid fa-user-slash",
                    "value": customers_with_debt,
                },
                {
                    "title": "Recaudación del Mes",
                    "icono" : "fa-solid fa-chart-line",          
                    "value": f"S/ {monthly_revenue}",
                },
                {
                    "title": "Recaudación del Día",
                    "icono" : "fa-solid fa-calendar-day",  
                    "value": f"S/ {daily_revenue}",
                },
            ],

            "top_debtors": top_debtors,
            "weekly_income": weekly_income,
            "charts": {
                "weekly_income": weekly_income,
                "debt_by_zone": debt_by_zone
            },
            "financial_comparison": {
                "debt": float(total_debt),
                "collected": float(total_collected)
            },
        })

class ProcessPayment(TenantSafeMixin, APIView):

    authentication_classes = []
    permission_classes = []

    def post(self, request):

        sdk = mercadopago.SDK(settings.MP_ACCESS_TOKEN)

        payment_data = {
            "transaction_amount": float(request.data["transaction_amount"]),
            "installments": int(request.data["installments"]),  # normalmente 1
            "token": request.data["token"],
            "payment_method_id": request.data["payment_method_id"],
            "issuer_id": request.data.get("issuer_id"),
            "description": "Pago servicio de agua",
            "payer": {
                "email": request.data["payer"]["email"],
                "identification": request.data["payer"].get("identification")
            },
            "additional_info": {
                "items": [
                    {
                        "id": "SERV-AGUA",
                        "title": "Servicio de agua potable",
                        "category_id": "utilities",
                        "quantity": 1,
                        "unit_price": float(request.data["transaction_amount"])
                    }
                ]
            },
            "external_reference": request.data["external_reference"],
            "metadata": {
                "tenant": request.data["tenant_schema"],
                "customer_id": request.data["customer_id"],
                "debt_ids": request.data["debt_ids"],
            }

        }

        payment_response = sdk.payment().create(payment_data)
        payment = payment_response["response"]

        return Response(payment)
    
class ProcessPaymentYape(TenantSafeMixin, APIView):

    authentication_classes = []
    permission_classes = []

    def post(self, request):

        sdk = mercadopago.SDK(settings.MP_ACCESS_TOKEN)
        token = request.data.get("token")
        amount = request.data.get("amount")

        request_options = mercadopago.config.RequestOptions()
        request_options.custom_headers = {
            "X-Idempotency-Key": str(uuid.uuid4())
        }

        payment_data = {
            "description": "Pago servicio agua",
            "payment_method_id": "yape",
            "token": token,
            "transaction_amount": float(amount),
            "installments": 1,
            "notification_url": settings.MP_WEBHOOK_URL,
            "external_reference": request.data["external_reference"],
            "metadata": {
                "tenant": request.data["tenant_schema"],
                "customer_id": request.data["customer_id"],
                "debt_ids": request.data["debt_ids"],
            },
            "payer": {
                "email": "pablo_joseph01@hotmail.com"
            },
        }

        payment_response = sdk.payment().create(
            payment_data,
            request_options
        )

        return Response(payment_response["response"])
    
class PaymentStatusView(TenantSafeMixin, APIView):
    
    authentication_classes = []
    permission_classes = []

    def get(self, request, payment_id):

        pay = Pay.objects.filter(payment_id=payment_id).first()

        if not pay:

            return Response({"status": "not_found"}, status=404)

        data = {
            "status": pay.status,
            "processed": pay.processed
        }

        if pay.processed:
        
            invoice = Invoice.objects.filter(payment_reference=pay.payment_id).first()

            if invoice:

                data.update({

                    "id": invoice.id,
                    "code" : invoice.code

                })

        return Response(data)

class DebtRefinancingViewSet(TenantSafeMixin, viewsets.ModelViewSet):

    queryset = DebtRefinancing.objects.all().select_related('customer').prefetch_related('installment_details')
    serializer_class = DebtRefinancingSerializer
    filter_backends = [DjangoFilterBackend,filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['customer','paid']
    search_fields = ['customer__full_name']
    ordering = ['-id']

    @action(detail=True, methods=['get'])
    def report(self, request, pk=None):

        tenant = request.tenant.schema_name
        company = Company.objects.first()
        logo_url = request.build_absolute_uri(settings.MEDIA_URL + f'{company.ruc}.jpeg')
        refinancing = self.get_object()

        template = None

        if tenant == 'chilca':

            template = 'refinancing/report.html'

        else:

            if refinancing.type == 'service':

                template = 'refinancing/pangoa_service.html'

            elif refinancing.type == 'debt':

                template = 'refinancing/pangoa_debt.html'

        installments = refinancing.installment_details.all()
        html_string = render_to_string(template,

            {
                'ref': refinancing,
                'installments': installments,
                'logo_url': logo_url
            }

        )

        print({
                'ref': refinancing,
                'installments': installments,
                'logo_url': logo_url
            })

        html = HTML(
            string=html_string,
            base_url=request.build_absolute_uri('/')
        )

        pdf_file = html.write_pdf()

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename=refinanciamiento_{refinancing.id}.pdf'

        return response

    @action(detail=False, methods=['get'])
    def resumen(self, request):

        refinancings = self.get_queryset()

        total_refinanced = refinancings.aggregate(
            total=Sum('total_amount_with_interest')
        )['total'] or 0

        total_pending = RefinancingInstallment.objects.filter(
            paid=False
        ).aggregate(
            total=Sum('total_amount')
        )['total'] or 0

        total_paid = RefinancingInstallment.objects.filter(
            paid=True
        ).aggregate(
            total=Sum('total_amount')
        )['total'] or 0

        total_installments_pending = (
            RefinancingInstallment.objects.filter(
                paid=False
            ).count()
        )

        total_installments_paid = (
            RefinancingInstallment.objects.filter(
                paid=True
            ).count()
        )

        return Response({

            "total_refinancings": refinancings.count(),

            "total_refinanced": total_refinanced,

            "total_paid": total_paid,

            "total_pending": total_pending,

            "installments_paid":
                total_installments_paid,

            "installments_pending":
                total_installments_pending

        })


        installment_id = request.data.get(
            'installment_id'
        )

        installment = get_object_or_404(

            RefinancingInstallment,

            id=installment_id,

            refinancing_id=pk

        )

        installment.paid = True

        installment.save()

        ####################################################
        # VALIDAR SI TODO ESTA PAGADO
        ####################################################

        refinancing = installment.refinancing

        pending = refinancing.installments.filter(
            paid=False
        ).exists()

        if not pending:

            refinancing.paid = True

            refinancing.save()

        return Response({

            "message": "Cuota pagada correctamente"

        })

class AtypicalConsumptionReportExcelView(TenantSafeMixin, APIView):

    HIGH_FACTOR = Decimal("1.8")
    LOW_FACTOR = Decimal("0.4")

    def get(self, request):

        period = request.GET.get("period")

        if not period:

            return Response(
                {"error": "Debe enviar el periodo"},
                status=400
            )

        zona = request.GET.get("zona")
        category = request.GET.get("category")

        queryset = (
            Reading.objects
            .filter(
                period=period,
                customer__state="active",
                customer__status=True,
            )
            .select_related(
                "customer",
                "customer__zona",
                "customer__category"
            )
            .order_by(
                "customer__zona__name",
                "customer__full_name"
            )
        )
        
        # =====================================================
        # CLIENTES NO LECTURADOS
        # =====================================================

        read_customer_ids = queryset.values_list(
            "customer_id",
            flat=True
        )

        not_read_queryset = Customer.objects.filter(
            has_meter=True,
            status=True,
            state="active",
        ).exclude(
            id__in=read_customer_ids
        )

        if zona:

            queryset = queryset.filter(customer__zona_id=zona)

        if category:

            queryset = queryset.filter(customer__category_id=category)


        # =====================================================
        # CONTENEDORES
        # =====================================================

        high_consumption = []
        low_consumption = []
        zero_consumption = []
        abrupt_variation = []
        suspicious_leaks = []
        suspended_with_consumption = []
        not_read_customers = []
        # =====================================================
        # ANALISIS
        # =====================================================

        for reading in queryset:

            customer = reading.customer

            historical = (
                Reading.objects
                .filter(
                    customer=customer,
                    period__lt=reading.period
                )
                .exclude(consumption__isnull=True)
                .order_by("-period")[:6]
            )

            avg_consumption = (
                historical.aggregate(
                    avg=Avg("consumption")
                )["avg"]
                or Decimal("0")
            )

            current = reading.consumption or Decimal("0")

            variation_percent = Decimal("0")

            if avg_consumption > 0:

                variation_percent = (
                    (
                        current - avg_consumption
                    ) / avg_consumption
                ) * 100

            # =================================================
            # DEUDA
            # =================================================

            has_debt = Debt.objects.filter(
                customer=customer,
                paid=False
            ).exists()

            base_data = {

                "codigo": customer.codigo,
                "tiene_medidor": "SI" if customer.has_meter else "NO",
                "cliente": customer.full_name,
                "zona": (
                    customer.zona.name
                    if customer.zona else "-"
                ),
                "categoria": (
                    customer.category.name
                    if customer.category else "-"
                ),
                "consumo_actual": float(current),
                "promedio": float(avg_consumption),
                "variacion": round(float(variation_percent), 2),
                "deuda": "SI" if has_debt else "NO",
                "estado": customer.state,
                "direccion": customer.address or "",
            }

            # =================================================
            # ALTO CONSUMO
            # =================================================

            if (
                avg_consumption > 0
                and current > avg_consumption * self.HIGH_FACTOR
            ):

                high_consumption.append(base_data)

            # =================================================
            # BAJO CONSUMO
            # =================================================

            elif (
                avg_consumption > 0
                and current < avg_consumption * self.LOW_FACTOR
            ):

                low_consumption.append(base_data)

            # =================================================
            # CONSUMO CERO
            # =================================================

            if current == 0:

                zero_consumption.append(base_data)

            # =================================================
            # VARIACION BRUSCA
            # =================================================

            if abs(variation_percent) >= 100:

                abrupt_variation.append(base_data)

            # =================================================
            # SUSPENDIDOS CON CONSUMO
            # =================================================

            if (customer.state in ["cut"]):

                suspended_with_consumption.append(base_data)

            # =================================================
            # POSIBLE FUGA
            # =================================================

            last_3 = list(
                Reading.objects.filter(
                    customer=customer,
                    period__lt=reading.period
                )
                .order_by("-period")[:3]
            )

            if len(last_3) == 3:

                consumptions = [
                    r.consumption for r in reversed(last_3)
                ]

                consumptions.append(current)

                if (
                    consumptions[0]
                    < consumptions[1]
                    < consumptions[2]
                    < consumptions[3]
                ):

                    suspicious_leaks.append(base_data)


        for customer in not_read_queryset:

            has_debt = Debt.objects.filter(
                customer=customer,
                paid=False
            ).exists()

            not_read_customers.append({

                "codigo": customer.codigo,

                "cliente": customer.full_name,

                "zona": (
                    customer.zona.name
                    if customer.zona else "-"
                ),

                "tiene_medidor": (
                    "SI" if customer.has_meter else "NO"
                ),

                "categoria": (
                    customer.category.name
                    if customer.category else "-"
                ),

                "consumo_actual": "",

                "promedio": "",

                "variacion": "",

                "deuda": "SI" if has_debt else "NO",

                "estado": customer.state,

                "direccion": customer.address or "",

            })
        # =====================================================
        # CREAR EXCEL
        # =====================================================

        wb = Workbook()

        # =====================================================
        # ESTILOS
        # =====================================================

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        red_fill = PatternFill(
            start_color="FFC7CE",
            end_color="FFC7CE",
            fill_type="solid"
        )

        yellow_fill = PatternFill(
            start_color="FFF3CD",
            end_color="FFF3CD",
            fill_type="solid"
        )

        blue_fill = PatternFill(
            start_color="D9EAF7",
            end_color="D9EAF7",
            fill_type="solid"
        )

        white_font = Font(
            bold=True,
            color="FFFFFF"
        )

        # =====================================================
        # FUNCION CREAR SHEET
        # =====================================================

        def create_sheet(title, data, fill=None):

            ws = wb.create_sheet(title)

            headers = [
                "CODIGO",
                "CLIENTE",
                "ZONA",
                "TIENE_MEDIDOR",
                "CATEGORIA",
                "CONSUMO ACTUAL",
                "PROMEDIO",
                "VARIACION %",
                "DEUDA",
                "ESTADO",
                "DIRECCION",
            ]

            ws.append(headers)

            # HEADER STYLE
            for cell in ws[1]:

                cell.font = white_font
                cell.fill = header_fill

            # DATA
            for item in data:

                row = [
                    item["codigo"],
                    item["cliente"],
                    item["zona"],
                    item["tiene_medidor"],
                    item["categoria"],
                    item["consumo_actual"],
                    item["promedio"],
                    item["variacion"],
                    item["deuda"],
                    item["estado"],
                    item["direccion"],
                ]

                ws.append(row)

            # COLOR FILAS
            if fill:

                for row in ws.iter_rows(
                    min_row=2,
                    max_row=ws.max_row
                ):

                    for cell in row:

                        cell.fill = fill

            # AUTOFILTER
            ws.auto_filter.ref = ws.dimensions

            # FREEZE
            ws.freeze_panes = "A2"

            # AUTO WIDTH
            for column_cells in ws.columns:

                length = max(
                    len(str(cell.value or ""))
                    for cell in column_cells
                )

                column_letter = get_column_letter(
                    column_cells[0].column
                )

                ws.column_dimensions[
                    column_letter
                ].width = length + 5

            return ws

        # =====================================================
        # ELIMINAR HOJA DEFAULT
        # =====================================================

        wb.remove(wb.active)

        # =====================================================
        # RESUMEN
        # =====================================================

        summary = wb.create_sheet("Resumen")

        summary.append(["INDICADOR", "TOTAL"])

        summary.append([
            "Clientes analizados",
            queryset.count()
        ])

        summary.append([
            "Alto consumo",
            len(high_consumption)
        ])

        summary.append([
            "Bajo consumo",
            len(low_consumption)
        ])

        summary.append([
            "Consumo cero",
            len(zero_consumption)
        ])

        summary.append([
            "Variación brusca",
            len(abrupt_variation)
        ])

        summary.append([
            "Posible fuga",
            len(suspicious_leaks)
        ])

        summary.append([
            "Clientes no lecturados",
            len(not_read_customers)
        ])


        for cell in summary[1]:

            cell.font = white_font
            cell.fill = header_fill

        summary.freeze_panes = "A2"

        # =====================================================
        # SHEETS
        # =====================================================

        create_sheet(
            "Alto Consumo",
            high_consumption,
            red_fill
        )

        create_sheet(
            "Bajo Consumo",
            low_consumption,
            yellow_fill
        )

        create_sheet(
            "Consumo Cero",
            zero_consumption,
            blue_fill
        )

        create_sheet(
            "Variacion Brusca",
            abrupt_variation,
            yellow_fill
        )

        create_sheet(
            "Posibles Fugas",
            suspicious_leaks,
            red_fill
        )

        create_sheet(
            "No Lecturados",
            not_read_customers,
            yellow_fill
        )


        # =====================================================
        # RESPONSE
        # =====================================================

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            )
        )

        filename = (
            f"reporte_consumos_atipicos_{period}.xlsx"
        )

        response[
            "Content-Disposition"
        ] = f'attachment; filename="{filename}"'

        wb.save(response)

        return response